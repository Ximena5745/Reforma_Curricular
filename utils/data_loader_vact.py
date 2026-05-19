"""
utils/data_loader_vact.py
Carga y procesa la hoja 'Etapas' del archivo VACT (Fase 2).
Fuente: data/raw/CONTROL MAESTRO DE REFORMA CURRICULAR.xlsx
"""

from __future__ import annotations

import re
import warnings
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

ROOT = Path(__file__).parent.parent
DATA_PATH = ROOT / "data" / "raw" / "CONTROL MAESTRO DE REFORMA CURRICULAR.xlsx"
DATA_SOURCE_NAME = "Control Maestro Reforma Curricular"
TZ_BOGOTA = ZoneInfo("America/Bogota")


_CACHED_EXCEL_MODIFIED: datetime | None = None


def _parse_office_datetime(value: str) -> datetime | None:
    """Parsea fecha ISO de metadatos Office (UTC) a datetime con zona."""
    if not value or not str(value).strip():
        return None
    s = str(value).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt


def _excel_core_modified(path: Path) -> datetime | None:
    """Última modificación guardada en Excel (docProps/core.xml), no mtime del SO."""
    try:
        with zipfile.ZipFile(path, "r") as zf:
            if "docProps/core.xml" not in zf.namelist():
                return None
            root = ET.fromstring(zf.read("docProps/core.xml"))
        for el in root.iter():
            if el.tag.endswith("}modified") and el.text:
                parsed = _parse_office_datetime(el.text)
                if parsed is not None:
                    return parsed
        return None
    except (OSError, zipfile.BadZipFile, ET.ParseError):
        return None


def _refresh_excel_modified_cache() -> None:
    global _CACHED_EXCEL_MODIFIED
    _CACHED_EXCEL_MODIFIED = _excel_core_modified(DATA_PATH)


def get_raw_data_updated_label() -> str:
    """Fecha/hora de última guardada del Excel (metadatos internos), horario Bogotá."""
    if not DATA_PATH.is_file():
        return "Actualizado: —"
    dt = _CACHED_EXCEL_MODIFIED
    if dt is None:
        dt = _excel_core_modified(DATA_PATH)
    if dt is None:
        dt = datetime.fromtimestamp(DATA_PATH.stat().st_mtime, tz=TZ_BOGOTA)
    else:
        dt = dt.astimezone(TZ_BOGOTA)
    return dt.strftime("Actualizado: %d/%m/%Y %H:%M")


PHASE_ROW = 7   # fila 8 Excel: fases
AREA_ROW = 8    # fila 9 Excel: áreas / responsables (fallback)
HEADER_ROW = 10  # fila 11 Excel: encabezados actividades
DATA_START_ROW = 11  # fila 12 Excel: primer programa

_AREA_HINTS = (
    "decanatura",
    "currículo",
    "curriculo",
    "secretaría",
    "secretaria",
    "docencia",
    "financiera",
    "aseguramiento",
    "gerencia de operaciones",
    "convenios",
    "homologacion",
    "homologación",
    "banner",
    "mercado",
)

INFO_COLS = [
    "FACULTAD",
    "ESCUELA",
    "NOMBRE DEL PROGRAMA",
    "MODALIDAD",
    "NIVEL",
    "SEDE",
    "PERIODO DE IMPLEMENTACIÓN",
]

FAC_ABREV = {
    "Facultad de Sociedad, Cultura y Creatividad": "FSCC",
    "Facultad de Ingeniería, Diseño e Innovación": "FIDI",
    "Facultad de Negocios, Gestión y Sostenibilidad": "FNGS",
}

from utils.poli_theme import (
    ETAPA_CLR,
    FACULTAD_CLR,
    STATUS_CLR,
    color_for_pct,
)

FAC_COLORS = FACULTAD_CLR
FAC_ABREV_INV = {v: k for k, v in FAC_ABREV.items()}

STATUS_LABEL = {
    "done": "Finalizado",
    "inprog": "En proceso",
    "nostart": "Sin iniciar",
    "devuelto": "Devuelto",
    "info": "Informativo",
    "na": "No aplica",
}

STATUS_COLOR = STATUS_CLR

# Etapas canónicas (orden de visualización)
ETAPAS_ORDEN = [
    "Alistamiento Curricular",
    "Diseño Curricular",
    "Desarrollo Curricular",
    "Implementación Curricular",
]

ETAPA_PCT_COL = {
    "Alistamiento Curricular": "pct_alistamiento",
    "Diseño Curricular": "pct_diseno",
    "Desarrollo Curricular": "pct_desarrollo",
    "Implementación Curricular": "pct_implementacion",
}

ETAPA_SLUG = {
    "Alistamiento Curricular": "alistamiento",
    "Diseño Curricular": "diseno",
    "Desarrollo Curricular": "desarrollo",
    "Implementación Curricular": "implementacion",
}

ETAPA_HEADER_CLR = ETAPA_CLR


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _parse_responsable_cell(v) -> str:
    s = " ".join(str(v).strip().split())
    if not s or s.lower() in ("none", "nan", ""):
        return "—"
    return s


def _norm_phase_name(s: str) -> str:
    """Normaliza nombres de fase del Excel."""
    t = _norm(s)
    if "alistamiento" in t:
        return "Alistamiento Curricular"
    if "dise" in t and "curricular" in t:
        return "Diseño Curricular"
    if "desarr" in t or "desarrollo" in t:
        return "Desarrollo Curricular"
    if "implement" in t:
        return "Implementación Curricular"
    if "fases de la reforma" in t:
        return "Info Programa"
    return str(s).strip()


def homologar_nivel(val) -> str:
    v = str(val).strip().lower()
    if any(v.startswith(p) for p in ("profesional", "tecnico", "técnico", "tecnologico", "tecnológico", "tecn")):
        return "Pregrado"
    if any(v.startswith(p) for p in ("maestria", "maestría", "especializacion", "especialización", "maest", "espec")):
        return "Posgrado"
    return ""


def _cls_status(v) -> str:
    s = " ".join(str(v).strip().lower().split())
    if not s or s in ("none", "nan", ""):
        return "na"
    if "no aplica" in s:
        return "na"
    if "devuelto" in s or "devuelta" in s:
        return "devuelto"
    if s in ("finalizado", "si", "sí"):
        return "done"
    if "sin iniciar" in s or s == "no":
        return "nostart"
    if s == "en proceso":
        return "inprog"
    if "aprobado" in s or "renovación" in s or "renovacion" in s:
        return "inprog"
    keywords_inprog = [
        "visita", "evaluación", "elaboración", "completitud", "resolución",
        "notificado", "pendiente",
    ]
    if any(k in s for k in keywords_inprog):
        return "inprog"
    if s == "informativo" or "informativo" in s:
        return "info"
    return "inprog" if s else "na"


def _cls_pct_value(v) -> float | None:
    """Convierte valor de celda a porcentaje 0-100."""
    s = str(v).strip().lower()
    if not s or s in ("none", "nan", "no aplica", "—", ""):
        return None
    try:
        f = float(s.replace("%", "").replace(",", "."))
        return round(f * 100 if 0 <= f <= 1 else f, 1)
    except Exception:
        return None


def _activity_score(cl: str) -> float | None:
    if cl == "done":
        return 100.0
    if cl == "inprog":
        return 50.0
    if cl == "nostart":
        return 0.0
    if cl == "devuelto":
        return 25.0
    return None  # na / info no cuentan


def _detect_area_row(raw: pd.DataFrame) -> int:
    """Detecta fila de áreas/responsables por palabras clave (filas 7–12 Excel)."""
    best_row = AREA_ROW
    best_score = 0
    for i in range(6, min(12, len(raw))):
        score = 0
        for j in range(raw.shape[1]):
            t = _norm(str(raw.iloc[i, j]))
            if t and t not in ("none", "nan") and any(h in t for h in _AREA_HINTS):
                score += 1
        if score > best_score:
            best_score = score
            best_row = i
    return best_row if best_score > 0 else AREA_ROW


def _forward_fill_area_values(raw_vals: list[str]) -> dict[int, str]:
    """Propaga el último responsable no vacío a columnas siguientes."""
    area_by_col: dict[int, str] = {}
    current = ""
    for j, s in enumerate(raw_vals):
        if s and s.lower() not in ("none", "nan"):
            current = s
        area_by_col[j] = _parse_responsable_cell(current) if current else "—"
    return area_by_col


def _build_area_by_col_openpyxl(area_row_idx: int, n_cols: int) -> dict[int, str]:
    """Lee fila de áreas resolviendo celdas combinadas con openpyxl."""
    if not DATA_PATH.is_file() or n_cols <= 0:
        return {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(DATA_PATH, data_only=True, read_only=False)
        if "Etapas" not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb["Etapas"]
        excel_row = area_row_idx + 1
        ncol = max(n_cols, ws.max_column or 0)
        raw_vals = [""] * ncol

        for c in range(1, ncol + 1):
            v = ws.cell(row=excel_row, column=c).value
            if v is not None and str(v).strip():
                raw_vals[c - 1] = " ".join(str(v).strip().split())

        for mr in ws.merged_cells.ranges:
            if mr.min_row <= excel_row <= mr.max_row:
                tl = ws.cell(row=mr.min_row, column=mr.min_col).value
                if tl is not None and str(tl).strip():
                    val = " ".join(str(tl).strip().split())
                    for c in range(mr.min_col, mr.max_col + 1):
                        raw_vals[c - 1] = val
        wb.close()
        return _forward_fill_area_values(raw_vals)
    except Exception:
        return {}


def _build_area_by_col(raw: pd.DataFrame) -> dict[int, str]:
    """Responsables por columna: openpyxl (merges) + fallback pandas + forward-fill."""
    if len(raw) == 0:
        return {}
    area_row_idx = _detect_area_row(raw)
    n_cols = raw.shape[1]

    area_by_col = _build_area_by_col_openpyxl(area_row_idx, n_cols)
    if area_by_col and sum(1 for v in area_by_col.values() if v != "—") >= 3:
        return area_by_col

    if len(raw) <= area_row_idx:
        return area_by_col
    area_row = raw.iloc[area_row_idx]
    raw_vals = []
    for j in range(n_cols):
        v = area_row.iloc[j] if j < len(area_row) else ""
        s = " ".join(str(v).strip().split()) if pd.notna(v) and str(v).strip() else ""
        raw_vals.append(s)
    return _forward_fill_area_values(raw_vals)


def _build_phase_column_map(raw: pd.DataFrame) -> tuple[dict, list[dict]]:
    """
    Retorna (phase_by_col_index, activities_list).
    activities_list: [{phase, name, col_idx, is_pct}, ...]
    """
    phase_row = raw.iloc[PHASE_ROW]
    area_by_col = _build_area_by_col(raw)
    header_row = raw.iloc[HEADER_ROW]

    phase_by_col: dict[int, str] = {}
    current_phase = "Info Programa"
    for j in range(raw.shape[1]):
        pv = phase_row.iloc[j]
        if pd.notna(pv) and str(pv).strip():
            current_phase = _norm_phase_name(str(pv))
        phase_by_col[j] = current_phase

    activities: list[dict] = []
    for j in range(raw.shape[1]):
        hv = header_row.iloc[j]
        if pd.isna(hv) or not str(hv).strip():
            continue
        hname = str(hv).strip()
        phase = phase_by_col.get(j, "Info Programa")

        is_pct = "% de avance" in _norm(hname) or hname.startswith("%")
        is_general = "general de reforma" in _norm(hname)

        if phase == "Info Programa" and not is_pct:
            continue

        if is_general:
            activities.append({
                "phase": "General",
                "name": hname,
                "col_idx": j,
                "is_pct": True,
                "is_general": True,
            })
            continue

        if is_pct:
            activities.append({
                "phase": phase,
                "name": hname,
                "col_idx": j,
                "is_pct": True,
                "is_general": False,
            })
        elif phase in ETAPAS_ORDEN:
            resp = area_by_col.get(j, "—")
            activities.append({
                "phase": phase,
                "name": hname,
                "col_idx": j,
                "is_pct": False,
                "is_general": False,
                "responsable": resp,
            })

    seen_act: dict[tuple[str, str], int] = {}
    for act in activities:
        if act.get("is_pct") or act["phase"] not in ETAPAS_ORDEN:
            continue
        key = (act["phase"], act["name"].strip().lower())
        if key in seen_act:
            warnings.warn(
                f"Actividad duplicada en Excel: «{act['name']}» ({act['phase']})",
                UserWarning,
                stacklevel=2,
            )
        seen_act[key] = seen_act.get(key, 0) + 1

    return phase_by_col, activities


def _build_activities_meta_list(activities_meta: list[dict]) -> list[dict]:
    """Metadatos de actividades con responsable desde el mapa de columnas."""
    built: list[dict] = []
    act_idx = 0
    for m in activities_meta:
        if m.get("is_pct") or m.get("is_general"):
            continue
        if m.get("phase") not in ETAPAS_ORDEN:
            continue
        built.append({
            "idx": act_idx,
            "phase": m["phase"],
            "name": m["name"],
            "responsable": m.get("responsable", "—"),
        })
        act_idx += 1
    return built


def _build_etapas_df() -> pd.DataFrame:
    global _ACTIVITIES_META

    _refresh_excel_modified_cache()
    raw = pd.read_excel(DATA_PATH, sheet_name="Etapas", header=None, dtype=str)
    raw = raw.fillna("")

    _, activities_meta = _build_phase_column_map(raw)
    header_row = raw.iloc[HEADER_ROW]

    # Índices fijos de columnas info (según fila 11 del Excel)
    info_idx = {
        "FACULTAD": 2,
        "ESCUELA": 3,
        "NOMBRE DEL PROGRAMA": 4,
        "MODALIDAD": 5,
        "NIVEL": 6,
        "SEDE": 8,
        "PERIODO DE IMPLEMENTACIÓN": 10,
    }

    data_raw = raw.iloc[DATA_START_ROW:].reset_index(drop=True)
    n_rows = len(data_raw)

    prog_idx = info_idx["NOMBRE DEL PROGRAMA"]
    mask = data_raw.iloc[:, prog_idx].astype(str).str.strip() != ""
    data_raw = data_raw[mask].reset_index(drop=True)
    n_rows = len(data_raw)

    df = pd.DataFrame()
    for col_name, idx in info_idx.items():
        if idx < data_raw.shape[1]:
            df[col_name] = data_raw.iloc[:, idx].astype(str).str.strip().values

    # Normalizar periodo
    if "PERIODO DE IMPLEMENTACIÓN" in df.columns:
        df["PERIODO DE IMPLEMENTACIÓN"] = (
            df["PERIODO DE IMPLEMENTACIÓN"].astype(str).str.split("\n").str[0].str.strip()
        )

    # Nivel homologado
    if "NIVEL" in df.columns:
        df["NIVEL_HOMOLOGADO"] = df["NIVEL"].apply(homologar_nivel)
    else:
        df["NIVEL_HOMOLOGADO"] = ""

    # Facultad abreviada
    if "FACULTAD" in df.columns:
        df["FACULTAD_ABREV"] = df["FACULTAD"].map(FAC_ABREV).fillna("—")
        df["FACULTAD_COLOR"] = df["FACULTAD_ABREV"].map(FAC_COLORS).fillna("#6e7681")
    else:
        df["FACULTAD_ABREV"] = "—"
        df["FACULTAD_COLOR"] = "#6e7681"

    # Procesar actividades por etapa (por índice de columna, evita duplicados de nombre)
    act_idx = 0
    for meta in activities_meta:
        col_idx = meta["col_idx"]
        series = data_raw.iloc[:n_rows, col_idx].astype(str).str.strip()
        series.index = range(n_rows)

        if meta.get("is_general"):
            df["avance_general_vact"] = series.apply(
                lambda v: _cls_pct_value(v) if _cls_pct_value(v) is not None else 0.0
            )
            continue

        if meta["is_pct"]:
            pct_key = ETAPA_PCT_COL.get(meta["phase"])
            if pct_key:
                df[pct_key] = series.apply(
                    lambda v: _cls_pct_value(v) if _cls_pct_value(v) is not None else np.nan
                )
            continue

        # Actividad individual
        cl_col = f"cl_act_{act_idx}"
        val_col = f"val_act_{act_idx}"
        df[cl_col] = series.apply(_cls_status)
        df[val_col] = series.apply(
            lambda v: str(v).strip() if str(v).strip() not in ("", "None", "nan") else "—"
        )
        df[f"act_phase_{act_idx}"] = meta["phase"]
        df[f"act_name_{act_idx}"] = meta["name"]
        df[f"act_owner_{act_idx}"] = meta.get("responsable", "—")
        act_idx += 1

    df["_n_activities"] = act_idx

    # Calcular % por etapa si no viene del Excel o para validar
    for etapa in ETAPAS_ORDEN:
        pct_key = ETAPA_PCT_COL[etapa]
        if pct_key not in df.columns:
            df[pct_key] = np.nan

        # Recalcular desde actividades si faltan valores
        mask_missing = df[pct_key].isna()
        if mask_missing.any():
            scores = []
            for i in range(act_idx):
                if df[f"act_phase_{i}"].iloc[0] != etapa:
                    continue
                cl_col = f"cl_act_{i}"
                row_scores = df[cl_col].apply(_activity_score)
                scores.append(row_scores)

            if scores:
                mat = pd.concat(scores, axis=1)
                calc = mat.mean(axis=1, skipna=True).round(1)
                df.loc[mask_missing, pct_key] = calc[mask_missing]

        # Rellenar NaN con 0
        df[pct_key] = df[pct_key].fillna(0).astype(float)

    if "avance_general_vact" not in df.columns:
        pct_cols = [ETAPA_PCT_COL[e] for e in ETAPAS_ORDEN]
        df["avance_general_vact"] = df[pct_cols].mean(axis=1).round(1)
    else:
        df["avance_general_vact"] = df["avance_general_vact"].fillna(0).astype(float)

    # Si avance general viene como decimal 0-1
    if df["avance_general_vact"].max() <= 1.0:
        df["avance_general_vact"] = (df["avance_general_vact"] * 100).round(1)

    for etapa in ETAPAS_ORDEN:
        pk = ETAPA_PCT_COL[etapa]
        if df[pk].max() <= 1.0:
            df[pk] = (df[pk] * 100).round(1)

    _ACTIVITIES_META = _build_activities_meta_list(activities_meta)
    return df


# Metadatos de actividades (se construyen al cargar)
_ACTIVITIES_META: list[dict] = []


def _ensure_activities_meta(df: pd.DataFrame) -> list[dict]:
    global _ACTIVITIES_META
    if _ACTIVITIES_META:
        return _ACTIVITIES_META
    n = int(df.get("_n_activities", pd.Series([0])).iloc[0]) if len(df) else 0
    meta = []
    for i in range(n):
        if f"act_phase_{i}" not in df.columns:
            break
        owner_col = f"act_owner_{i}"
        owner = df[owner_col].iloc[0] if owner_col in df.columns else "—"
        meta.append({
            "idx": i,
            "phase": df[f"act_phase_{i}"].iloc[0],
            "name": df[f"act_name_{i}"].iloc[0],
            "responsable": _parse_responsable_cell(owner),
        })
    _ACTIVITIES_META = meta
    return meta


# Alias para reglas de riesgo / alertas (nombres normalizados del Excel)
ACTIVITY_ALIASES: dict[str, list[str]] = {
    "proyecciones_formato_5": [
        "5.formato de proyecciones",
        "5. formato de proyecciones",
    ],
    "proyecciones_fin": [
        "5.formato de proyecciones",
        "formato de proyecciones académicas y financieras",
        "formato de proyecciones academicas y financieras",
        "estudio de viabilidad financiera",
    ],
    "produccion_contenido": [
        "producción del contenido",
        "produccion del contenido",
        "% avance contenidos",
        "contenidos virtuales",
    ],
    "banner_convenios": [
        "parametrización en banner",
        "parametrizacion en banner",
        "convenios y homologaciones",
    ],
    "syllabus": ["syllabus aprobado", "syllabus"],
    "convenios": ["formato maestro de convenios", "maestro de convenios"],
    "cronograma_men": [
        "cronograma de trámites frente al men",
        "cronograma de tramites frente al men",
        "trámites frente al men",
    ],
}

_ACTIVITY_IDX_CACHE: dict[str, int | None] = {}


def _resolve_activity_idx(df: pd.DataFrame, key: str) -> int | None:
    if key in _ACTIVITY_IDX_CACHE:
        return _ACTIVITY_IDX_CACHE[key]
    patterns = ACTIVITY_ALIASES.get(key, [])
    if not patterns:
        _ACTIVITY_IDX_CACHE[key] = None
        return None
    meta = _ensure_activities_meta(df)
    for m in meta:
        name_n = _norm(m["name"])
        for pat in patterns:
            if pat in name_n or name_n in pat:
                _ACTIVITY_IDX_CACHE[key] = m["idx"]
                return m["idx"]
    _ACTIVITY_IDX_CACHE[key] = None
    return None


def activity_cl(df: pd.DataFrame, key: str) -> pd.Series:
    """Serie de clasificación (done/inprog/...) para una actividad por alias."""
    idx = _resolve_activity_idx(df, key)
    if idx is None:
        return pd.Series("na", index=df.index)
    col = f"cl_act_{idx}"
    if col not in df.columns:
        return pd.Series("na", index=df.index)
    return df[col]


def activity_val(df: pd.DataFrame, key: str) -> pd.Series:
    """Serie de valor crudo de actividad por alias."""
    idx = _resolve_activity_idx(df, key)
    if idx is None:
        return pd.Series("—", index=df.index)
    col = f"val_act_{idx}"
    if col not in df.columns:
        return pd.Series("—", index=df.index)
    return df[col]


def activity_not_done(df: pd.DataFrame, key: str) -> pd.Series:
    """True si la actividad no está finalizada."""
    cl = activity_cl(df, key)
    return cl != "done"


def activity_in_progress(df: pd.DataFrame, key: str) -> pd.Series:
    cl = activity_cl(df, key)
    return cl == "inprog"


def activity_status_is(df: pd.DataFrame, key: str, status: str) -> pd.Series:
    """True si la actividad (por alias) tiene la clasificación indicada."""
    return activity_cl(df, key) == status


def activity_val_contains(df: pd.DataFrame, key: str, substring: str) -> pd.Series:
    vals = activity_val(df, key).astype(str).str.lower()
    return vals.str.contains(substring.lower(), na=False)


def load_etapas_data() -> pd.DataFrame:
    """Retorna DataFrame procesado de la hoja Etapas."""
    global _ACTIVITIES_META, _ACTIVITY_IDX_CACHE
    _ACTIVITY_IDX_CACHE = {}
    try:
        import streamlit as st

        mtime = int(DATA_PATH.stat().st_mtime) if DATA_PATH.is_file() else 0

        @st.cache_data
        def _cached(_v: int, _mtime: int):
            return _build_etapas_df()

        df = _cached(4, mtime)
    except Exception:
        df = _build_etapas_df()
    if not _ACTIVITIES_META:
        _ensure_activities_meta(df)
    return df


def apply_filters_vact(
    df: pd.DataFrame,
    modalidad=None,
    facultad=None,
    periodo=None,
    nivel=None,
) -> pd.DataFrame:
    """Filtros para Fase 2. facultad acepta nombres completos o abreviaturas."""

    def _to_list(v):
        if not v:
            return []
        return list(v) if not isinstance(v, str) else [v]

    mods = _to_list(modalidad)
    facs = _to_list(facultad)
    pers = _to_list(periodo)
    nivs = _to_list(nivel)

    out = df.copy()
    if mods and "MODALIDAD" in out.columns:
        out = out[out["MODALIDAD"].isin(mods)]
    if facs and "FACULTAD" in out.columns:
        full_facs = [FAC_ABREV_INV.get(f, f) for f in facs]
        out = out[out["FACULTAD"].isin(full_facs)]
    if pers and "PERIODO DE IMPLEMENTACIÓN" in out.columns:
        mask = out["PERIODO DE IMPLEMENTACIÓN"].isin(pers)
        if "2027-1" in pers:
            mask = mask | out["PERIODO DE IMPLEMENTACIÓN"].str.contains("2027-1", na=False)
        if any("oferta" in str(p).lower() for p in pers):
            mask = mask | out["PERIODO DE IMPLEMENTACIÓN"].str.contains("oferta", case=False, na=False)
        out = out[mask]
    if nivs and "NIVEL_HOMOLOGADO" in out.columns:
        out = out[out["NIVEL_HOMOLOGADO"].isin(nivs)]
    return out.reset_index(drop=True)


def get_etapas_by_programa(df: pd.DataFrame, nombre_programa: str) -> dict:
    """Actividades + estados de un programa."""
    meta = _ensure_activities_meta(df)
    row = df[df["NOMBRE DEL PROGRAMA"].astype(str).str.strip() == str(nombre_programa).strip()]
    if row.empty:
        return {"programa": nombre_programa, "etapas": {}}
    row = row.iloc[0]
    result: dict = {"programa": nombre_programa, "etapas": {}}
    for etapa in ETAPAS_ORDEN:
        acts = []
        for m in meta:
            if m["phase"] != etapa:
                continue
            i = m["idx"]
            cl = row.get(f"cl_act_{i}", "na")
            val = row.get(f"val_act_{i}", "—")
            score = _activity_score(cl)
            acts.append({
                "nombre": m["name"],
                "estado": STATUS_LABEL.get(cl, cl),
                "estado_key": cl,
                "valor": val,
                "pct": score if score is not None else "—",
                "responsable": m.get("responsable", "—"),
            })
        pct_col = ETAPA_PCT_COL[etapa]
        result["etapas"][etapa] = {
            "pct": float(row.get(pct_col, 0) or 0),
            "actividades": acts,
        }
    result["avance_general"] = float(row.get("avance_general_vact", 0) or 0)
    result["info"] = {
        "FACULTAD": row.get("FACULTAD", "—"),
        "FACULTAD_ABREV": row.get("FACULTAD_ABREV", "—"),
        "FACULTAD_COLOR": row.get("FACULTAD_COLOR", "#6e7681"),
        "ESCUELA": row.get("ESCUELA", "—"),
        "MODALIDAD": row.get("MODALIDAD", "—"),
        "NIVEL": row.get("NIVEL", "—"),
        "NIVEL_HOMOLOGADO": row.get("NIVEL_HOMOLOGADO", "—"),
        "SEDE": row.get("SEDE", "—"),
        "PERIODO DE IMPLEMENTACIÓN": row.get("PERIODO DE IMPLEMENTACIÓN", "—"),
    }
    return result


def get_estadisticas_etapa(df: pd.DataFrame, etapa_name: str) -> dict:
    """Estadísticas agregadas de una etapa sobre el df filtrado."""
    pct_col = ETAPA_PCT_COL.get(etapa_name)
    if not pct_col or pct_col not in df.columns or len(df) == 0:
        return {
            "pct_promedio": 0,
            "done": 0,
            "inprog": 0,
            "devuelto": 0,
            "nostart": 0,
            "info": 0,
            "na": 0,
            "total_act": 0,
            "n_programas": 0,
        }

    meta = _ensure_activities_meta(df)
    acts_meta = [m for m in meta if m["phase"] == etapa_name]
    done = inprog = devuelto = nostart = info = na = 0
    for m in acts_meta:
        col = f"cl_act_{m['idx']}"
        if col not in df.columns:
            continue
        for cl in df[col]:
            if cl == "done":
                done += 1
            elif cl == "inprog":
                inprog += 1
            elif cl == "devuelto":
                devuelto += 1
            elif cl == "nostart":
                nostart += 1
            elif cl == "info":
                info += 1
            else:
                na += 1

    return {
        "pct_promedio": round(float(df[pct_col].mean()), 1),
        "done": done,
        "inprog": inprog,
        "devuelto": devuelto,
        "nostart": nostart,
        "info": info,
        "na": na,
        "total_act": len(acts_meta) * len(df) if acts_meta else 0,
        "n_programas": len(df),
    }


def get_detalle_etapa(df: pd.DataFrame, etapa_name: str) -> dict:
    """Desglose ampliado de una etapa: estados, % y lista de actividades."""
    stats = get_estadisticas_etapa(df, etapa_name)
    meta = _ensure_activities_meta(df)
    acts_meta = [m for m in meta if m["phase"] == etapa_name]
    n_prog = len(df)
    actividades = []
    for m in acts_meta:
        col = f"cl_act_{m['idx']}"
        if col not in df.columns:
            continue
        done = int((df[col] == "done").sum())
        inprog = int((df[col] == "inprog").sum())
        devuelto = int((df[col] == "devuelto").sum())
        nostart = int((df[col] == "nostart").sum())
        info = int((df[col] == "info").sum())
        na = int((df[col] == "na").sum())
        scores = df[col].apply(_activity_score)
        pct_avance = round(float(scores.dropna().mean()), 1) if scores.notna().any() else 0.0
        actividades.append({
            "nombre": m["name"],
            "done": done,
            "inprog": inprog,
            "devuelto": devuelto,
            "nostart": nostart,
            "info": info,
            "na": na,
            "pct_done": round(done / n_prog * 100, 1) if n_prog else 0,
            "pct_avance": pct_avance,
        })
    actividades.sort(key=lambda a: (-a["pct_done"], a["nombre"]))
    total_cells = stats.get("total_act") or 0
    pct_por_estado = {}
    for k in ("done", "inprog", "devuelto", "nostart", "info", "na"):
        pct_por_estado[k] = round(stats[k] / total_cells * 100, 1) if total_cells else 0
    return {
        **stats,
        "actividades": actividades,
        "pct_por_estado": pct_por_estado,
        "n_actividades": len(acts_meta),
    }


def count_actividades_completadas(row, etapa_name: str, df: pd.DataFrame) -> tuple[int, int]:
    """Retorna (completadas, total) para una fila y etapa."""
    meta = _ensure_activities_meta(df)
    acts = [m for m in meta if m["phase"] == etapa_name]
    total = len(acts)
    done = sum(1 for m in acts if row.get(f"cl_act_{m['idx']}") == "done")
    return done, total
