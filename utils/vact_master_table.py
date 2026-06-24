"""
Tabla maestra VACT, heatmap y exportación Excel.
"""
from __future__ import annotations

import io

import pandas as pd
import streamlit as st
import streamlit.components.v1 as html_comp

from utils.data_loader_vact import (
    ETAPAS_ORDEN,
    ETAPA_HEADER_CLR,
    ETAPA_PCT_COL,
    ETAPA_SLUG,
    STATUS_LABEL,
    _ensure_activities_meta,
    get_etapas_by_programa,
)
from utils.poli_theme import (
    BG_ROW,
    BG_ROW_ALT,
    BG_TABLE,
    BORDER_ROW,
    BORDER_TABLE,
    BRAND_PRIMARY,
    BRAND_HIGHLIGHT,
    MODALIDAD_CLR,
    PERIODO_CLR,
    STATUS_CLR,
    TEXT_LIGHT,
    TEXT_MUTED,
    TEXT_NA,
    TEXT_PRIMARY,
    TEXT_SUBTLE,
    ETAPA_CLR,
    badge_html,
    p_bar_html,
    pct_bar_colors,
    status_icon_html,
)

# Tinte suave en celdas de datos del Excel (variante B)
_EXCEL_STATUS_TINT = 0.10
_EXCEL_PCT_TINT = 0.08

_META_EXPORT = [
    ("NOMBRE DEL PROGRAMA", "Programa"),
    ("FACULTAD", "Facultad"),
    ("ESCUELA", "Escuela"),
    ("MODALIDAD", "Modalidad"),
    ("NIVEL", "Nivel"),
    ("SEDE", "Sede"),
    ("PERIODO DE IMPLEMENTACIÓN", "Período"),
]

_HDR_ROW1_H = 36


def _p_esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


_STANDARD_STATUS_VALS = {
    "finalizado", "si", "sí", "en proceso", "sin iniciar", "no aplica",
    "devuelto", "devuelta", "no", "informativo", "—", "", "nan", "none",
}


def _vact_act_icon(cl: str, val=None) -> str:
    lbl = STATUS_LABEL.get(cl, cl)
    val_s = str(val).strip() if val is not None else ""

    # Mostrar texto si el valor es descriptivo (no es un estado estándar)
    if val_s and val_s.lower() not in _STANDARD_STATUS_VALS and val_s != "—":
        return (
            f'<span style="font-size:10px;color:{TEXT_SUBTLE};display:inline-block;'
            f'max-width:130px;line-height:1.3;word-break:break-word;text-align:left;'
            f'white-space:normal" title="{_p_esc(val_s)}">{_p_esc(val_s[:60])}</span>'
        )
    if cl == "na":
        return f'<span style="color:{TEXT_NA};font-size:13px;font-weight:600" title="{_p_esc(lbl)}">N/A</span>'
    icon = status_icon_html(cl, size=16, title=lbl)
    return icon if icon else f'<span style="color:{TEXT_LIGHT}">—</span>'


def _status_legend_html() -> str:
    parts = []
    for cl in ("done", "inprog", "nostart", "devuelto", "info", "na"):
        lbl = STATUS_LABEL.get(cl, cl)
        if cl == "na":
            icon = f'<span style="color:{TEXT_NA};font-weight:600">N/A</span>'
        else:
            icon = status_icon_html(cl, size=14, title=lbl)
        parts.append(
            f'<span style="margin-right:16px;font-size:11px;color:{TEXT_MUTED}">{icon} {lbl}</span>'
        )
    return (
        '<div style="display:flex;flex-wrap:wrap;gap:4px;margin-top:8px;padding:8px 4px">'
        + "".join(parts)
        + f'<span style="font-size:10px;color:{TEXT_LIGHT};margin-left:8px">'
        "· Pulse + en cada etapa para ver actividades</span></div>"
    )


def _blend_with_white(hex_color: str, tint: float) -> str:
    h = hex_color.lstrip("#")
    if len(h) != 6:
        return BG_ROW
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    w = 1.0 - tint
    return (
        f"#{int(255 * w + r * tint):02x}"
        f"{int(255 * w + g * tint):02x}"
        f"{int(255 * w + b * tint):02x}"
    )


def _etapa_hdr_tints(clr: str) -> dict[str, str]:
    return {
        "etapa_bg": clr,
        "etapa_fg": "#FFFFFF",
        "act_bg": _blend_with_white(clr, 0.09),
        "act_fg": TEXT_PRIMARY,
        "pct_bg": _blend_with_white(clr, 0.17),
        "pct_fg": TEXT_PRIMARY,
        "cell_act": _blend_with_white(clr, 0.06),
        "cell_act_alt": _blend_with_white(clr, 0.11),
        "border": _blend_with_white(clr, 0.28),
    }


_MASTER_TABLE_CSS = f"""
<style>
html, body {{ margin: 0; padding: 0; height: 100%; background: {BG_TABLE}; }}
.vact-table-scroller {{
  height: 100%;
  overflow: auto;
  -webkit-overflow-scrolling: touch;
  border-radius: 12px;
  border: 1.5px solid {BORDER_TABLE};
  box-shadow: 0 2px 12px rgba(15,56,90,.10);
  background: {BG_TABLE};
}}
.vact-master-table {{
  width: 100%; table-layout: auto; border-collapse: separate; border-spacing: 0;
  font-family: 'Segoe UI', sans-serif; min-width: max-content;
}}
.vact-master-table thead tr.hdr-etapa th {{
  position: sticky;
  top: 0;
  z-index: 4;
  height: {_HDR_ROW1_H}px;
  padding: 5px 8px;
  border-bottom: none;
  background-clip: padding-box;
  vertical-align: middle;
}}
.vact-master-table thead tr.hdr-act th {{
  position: sticky;
  top: {_HDR_ROW1_H}px;
  z-index: 3;
  padding: 5px 5px;
  border-top: none;
  background-clip: padding-box;
  vertical-align: bottom;
  opacity: 1;
}}
.vact-master-table td[class^="col-act-"] {{
  opacity: 1;
}}
.vact-master-table thead th.th-pin-left {{
  left: 0;
  z-index: 6 !important;
  box-shadow: 2px 0 4px rgba(0,0,0,.06);
}}
.vact-master-table thead th.th-pin-gen {{
  z-index: 5 !important;
}}
</style>
"""


_TOGGLE_JS = """
<script>
function toggleEtapa(slug) {
  var cols = document.querySelectorAll('.col-act-' + slug);
  var hdr = document.getElementById('hdr-etapa-' + slug);
  var btn = document.getElementById('btn-etapa-' + slug);
  if (!hdr || !btn) return;
  var show = cols.length && cols[0].style.display === 'none';
  for (var i = 0; i < cols.length; i++) {
    cols[i].style.display = show ? 'table-cell' : 'none';
  }
  var expanded = parseInt(hdr.getAttribute('data-colspan-expanded') || '1', 10);
  hdr.colSpan = show ? expanded : 1;
  btn.textContent = show ? '\u2212' : '+';
}
</script>
"""


def _short_label(text: str, max_len: int = 28) -> str:
    t = str(text).strip()
    return t if len(t) <= max_len else t[: max_len - 1] + "…"


def master_activities_table_html(df: pd.DataFrame) -> str:
    meta_all = _ensure_activities_meta(df)
    etapa_blocks = []
    ncols = 6
    for etapa in ETAPAS_ORDEN:
        meta = [m for m in meta_all if m["phase"] == etapa]
        clr_etapa = ETAPA_HEADER_CLR.get(etapa, BRAND_PRIMARY)
        etapa_blocks.append({
            "etapa": etapa,
            "slug": ETAPA_SLUG[etapa],
            "meta": meta,
            "pct_col": ETAPA_PCT_COL[etapa],
            "tints": _etapa_hdr_tints(clr_etapa),
        })
        ncols += 1 + len(meta)
    ncols += 1

    TH = (
        f'style="background:{BRAND_PRIMARY};color:#FFFFFF;font-size:9px;font-weight:700;'
        'padding:6px 4px;text-align:center;white-space:nowrap;'
        'border-right:1px solid rgba(255,255,255,0.10)"'
    )
    TH_GEN = (
        f'class="th-pin-gen" style="background:{BRAND_PRIMARY};color:#FFFFFF;font-size:9px;font-weight:800;'
        'padding:8px 6px;text-align:center;white-space:normal;line-height:1.25;'
        f'border-left:2px solid {BRAND_HIGHLIGHT};min-width:88px"'
    )
    THL = (
        f'class="th-pin-left" style="background:{BRAND_PRIMARY};color:#FFFFFF;font-size:9px;font-weight:700;'
        'padding:6px 8px;text-align:left;white-space:nowrap;'
        'border-right:1px solid rgba(255,255,255,0.10);min-width:140px;max-width:200px"'
    )

    h1 = [
        f'<th rowspan="2" {THL}>Programa</th>',
        f'<th rowspan="2" {TH}>Facultad</th>',
        f'<th rowspan="2" {TH}>Modalidad</th>',
        f'<th rowspan="2" {TH}>Nivel</th>',
        f'<th rowspan="2" {TH}>Sede</th>',
        f'<th rowspan="2" {TH}>Período</th>',
    ]
    h2 = []
    for blk in etapa_blocks:
        slug, etapa, meta = blk["slug"], blk["etapa"], blk["meta"]
        t = blk["tints"]
        n_span_expanded = 1 + len(meta)
        short = etapa.replace(" Curricular", "")
        pct_lbl = f"% Av. {_short_label(short, 18)}"
        h1.append(
            f'<th id="hdr-etapa-{slug}" colspan="1" data-colspan-expanded="{n_span_expanded}"'
            f' class="hdr-etapa-{slug}"'
            f' style="background:{t["etapa_bg"]};color:{t["etapa_fg"]};'
            f'font-size:10px;font-weight:800;text-align:center;'
            f'border-right:1px solid rgba(255,255,255,0.15)">'
            f'<button type="button" id="btn-etapa-{slug}" onclick="toggleEtapa(\'{slug}\')" '
            f'style="cursor:pointer;border:1px solid rgba(255,255,255,0.55);background:rgba(255,255,255,0.15);'
            f'color:#fff;border-radius:4px;width:22px;height:22px;font-weight:800;margin-right:6px">+</button>'
            f'{_p_esc(short)}</th>'
        )
        th_act = (
            f'style="background:{t["act_bg"]};color:{t["act_fg"]};font-size:9px;font-weight:700;'
            f'text-align:center;white-space:normal;line-height:1.3;'
            f'word-break:break-word;overflow-wrap:anywhere;min-width:90px;max-width:180px;'
            f'border-right:1px solid {t["border"]};display:none"'
        )
        th_pct_etapa = (
            f'style="background:{t["pct_bg"]};color:{t["pct_fg"]};font-size:9px;font-weight:700;'
            f'text-align:center;white-space:normal;line-height:1.2;'
            f'border-right:1px solid {t["border"]};min-width:72px"'
        )
        for m in meta:
            h2.append(
                f'<th class="col-act-{slug}" {th_act} title="{_p_esc(m["name"])}">'
                f'{_p_esc(m["name"])}</th>'
            )
        h2.append(
            f'<th class="col-pct-{slug}" {th_pct_etapa} title="{_p_esc(etapa)}">{_p_esc(pct_lbl)}</th>'
        )

    h1.append(
        f'<th rowspan="2" {TH_GEN} title="Avance General de Reforma">'
        f'% Av.<br>General<br>Reforma</th>'
    )

    rows = []
    cur_per = None
    row_idx = 0
    sort_cols = [c for c in ("PERIODO DE IMPLEMENTACIÓN", "NOMBRE DEL PROGRAMA") if c in df.columns]
    df_sorted = df.sort_values(sort_cols) if sort_cols else df

    for _, row in df_sorted.iterrows():
        per = str(row.get("PERIODO DE IMPLEMENTACIÓN", "—")).strip()
        if per != cur_per:
            cur_per = per
            ph_clr = PERIODO_CLR.get(per, TEXT_LIGHT)
            per_cnt = int((df_sorted["PERIODO DE IMPLEMENTACIÓN"].astype(str).str.strip() == per).sum())
            rows.append(
                f'<tr><td colspan="{ncols}" style="background:{ph_clr};color:#FFFFFF;'
                f'font-size:11px;font-weight:800;padding:7px 12px;letter-spacing:.3px">'
                f'{_p_esc(per)} · {per_cnt} PROGRAMAS</td></tr>'
            )
            row_idx = 0
        rbg = BG_ROW if row_idx % 2 == 0 else BG_ROW_ALT
        row_idx += 1
        TD = (
            f'style="padding:4px 3px;text-align:center;vertical-align:middle;'
            f'border-bottom:1px solid {BORDER_ROW};background:{rbg}"'
        )
        TDL = (
            f'style="padding:4px 6px;text-align:left;vertical-align:middle;'
            f'border-bottom:1px solid {BORDER_ROW};background:{rbg};min-width:140px;max-width:200px;'
            f'position:sticky;left:0;z-index:1;box-shadow:2px 0 4px rgba(0,0,0,.04)"'
        )
        prog = _p_esc(row.get("NOMBRE DEL PROGRAMA", "—"))
        fac_s = row.get("FACULTAD_ABREV", "—")
        fac_c = row.get("FACULTAD_COLOR", "#9aabb5")
        mod = str(row.get("MODALIDAD", "—"))
        mod_c = MODALIDAD_CLR.get(mod, TEXT_LIGHT)
        per_c = PERIODO_CLR.get(per, TEXT_LIGHT)
        nivel = row.get("NIVEL_HOMOLOGADO", row.get("NIVEL", "—"))

        cells = [
            f'<td {TDL}><span style="font-size:11px;font-weight:600;color:{TEXT_PRIMARY}">{prog}</span></td>',
            f'<td {TD}><span style="font-size:10px;font-weight:700;color:{fac_c}">{_p_esc(fac_s)}</span></td>',
            f'<td {TD}>{badge_html(mod, mod_c)}</td>',
            f'<td {TD}><span style="font-size:10px;color:{TEXT_SUBTLE}">{_p_esc(nivel)}</span></td>',
            f'<td {TD}><span style="font-size:10px;color:{TEXT_SUBTLE}">{_p_esc(row.get("SEDE", "—"))}</span></td>',
            f'<td {TD}>{badge_html(per, per_c)}</td>',
        ]
        for blk in etapa_blocks:
            slug = blk["slug"]
            t = blk["tints"]
            pct = float(row.get(blk["pct_col"], 0) or 0)
            act_bg = t["cell_act_alt"] if row_idx % 2 == 0 else t["cell_act"]
            for m in blk["meta"]:
                cl = row.get(f"cl_act_{m['idx']}", "na")
                val = row.get(f"val_act_{m['idx']}", "—")
                cells.append(
                    f'<td class="col-act-{slug}" style="display:none;padding:4px 3px;text-align:center;'
                    f'vertical-align:middle;border-bottom:1px solid {BORDER_ROW};background:{act_bg}">'
                    f'{_vact_act_icon(cl, val)}</td>'
                )
            cells.append(
                f'<td class="col-pct-{slug}" style="padding:4px 3px;text-align:center;'
                f'vertical-align:middle;border-bottom:1px solid {BORDER_ROW};background:{rbg}">'
                f'{p_bar_html(pct)}</td>'
            )
        gen_pct = float(row.get("avance_general_vact", 0) or 0)
        cells.append(
            f'<td class="col-pct-general" {TD} style="border-left:2px solid {BORDER_ROW}">'
            f'{p_bar_html(gen_pct)}</td>'
        )
        rows.append("<tr>" + "".join(cells) + "</tr>")

    if not rows:
        return f"<p style='padding:20px;color:{TEXT_MUTED}'>No hay programas con los filtros seleccionados.</p>"

    return (
        _TOGGLE_JS
        + _MASTER_TABLE_CSS
        + '<div class="vact-table-scroller">'
        '<table class="vact-master-table">'
        '<thead><tr class="hdr-etapa">' + "".join(h1) + '</tr><tr class="hdr-act">' + "".join(h2) + "</tr></thead><tbody>"
        + "".join(rows)
        + "</tbody></table></div>"
        + _status_legend_html()
    )


def render_master_table(df: pd.DataFrame) -> None:
    if len(df) == 0:
        st.info("Sin programas para los filtros seleccionados.")
        return
    n_per = df["PERIODO DE IMPLEMENTACIÓN"].nunique() if "PERIODO DE IMPLEMENTACIÓN" in df.columns else 0
    h = min(120 + 42 * len(df) + 42 * n_per, 900)
    html_comp.html(master_activities_table_html(df), height=h, scrolling=False)


def render_master_table_by_periodo(df: pd.DataFrame, *, key_prefix: str = "master") -> None:
    """Tabla maestra con una pestaña por período de implementación."""
    if len(df) == 0:
        st.info("Sin programas para los filtros seleccionados.")
        return

    per_col = "PERIODO DE IMPLEMENTACIÓN"
    if per_col not in df.columns:
        render_master_table(df)
        return

    periodos = sorted(df[per_col].dropna().astype(str).str.strip().unique().tolist())
    if not periodos:
        render_master_table(df)
        return

    if len(periodos) == 1:
        mask = df[per_col].astype(str).str.strip() == periodos[0]
        render_master_table(df.loc[mask])
        return

    labels = [
        f"{p} ({int((df[per_col].astype(str).str.strip() == p).sum())})"
        for p in periodos
    ]
    tabs = st.tabs(labels)
    for tab, per in zip(tabs, periodos):
        with tab:
            mask = df[per_col].astype(str).str.strip() == per
            render_master_table(df.loc[mask])


def _pct_color(val: float) -> str:
    if val >= 80:
        return "#059669"
    if val >= 50:
        return "#2563eb"
    if val >= 30:
        return "#d97706"
    return "#dc2626"


def _hex6(color: str) -> str:
    h = str(color).lstrip("#")
    return h.upper() if len(h) == 6 else "9AABB5"


def _hex_fill(hex_color: str):
    from openpyxl.styles import PatternFill

    h = _hex6(hex_color)
    return PatternFill(start_color=h, end_color=h, fill_type="solid")


def _activity_display(row: pd.Series, act_idx: int) -> tuple[str, str]:
    cl = str(row.get(f"cl_act_{act_idx}", "na") or "na")
    val = row.get(f"val_act_{act_idx}", "—")
    if cl == "info" and str(val).strip() not in ("—", "", "nan", "None"):
        return str(val).strip(), cl
    return STATUS_LABEL.get(cl, cl), cl


def _status_font_color(cl: str) -> str:
    """Texto sobre fondo sólido — solo hoja Leyenda del Excel."""
    return "475569" if cl == "na" else "FFFFFF"


def _status_excel_style(cl: str) -> tuple[str, str]:
    """Fondo tintado + texto en color institucional (exportación detalle)."""
    if cl == "na":
        return _hex6(BG_ROW), _hex6(TEXT_NA)
    solid = STATUS_CLR.get(cl, "9aabb5")
    return _hex6(_blend_with_white(solid, _EXCEL_STATUS_TINT)), _hex6(solid)


def _pct_excel_style(pct: float) -> tuple[str, str]:
    """Fondo tintado suave + texto según escala pct_bar_colors."""
    fg, bar = pct_bar_colors(pct)
    return _hex6(_blend_with_white(bar, _EXCEL_PCT_TINT)), _hex6(fg)


def _excel_row_bg(ri: int, data_start: int) -> str:
    return _hex6(BG_ROW if (ri - data_start) % 2 == 0 else BG_ROW_ALT)


def _build_export_specs(df: pd.DataFrame) -> list[dict]:
    specs: list[dict] = [
        {"kind": "meta", "field": field, "label": label} for field, label in _META_EXPORT
    ]
    meta_all = _ensure_activities_meta(df)
    for etapa in ETAPAS_ORDEN:
        acts = [m for m in meta_all if m["phase"] == etapa]
        short = etapa.replace(" Curricular", "")
        for m in acts:
            specs.append({
                "kind": "act",
                "idx": m["idx"],
                "label": m["name"],
                "etapa": etapa,
            })
        specs.append({
            "kind": "pct",
            "field": ETAPA_PCT_COL[etapa],
            "label": f"% Av. {short}",
            "etapa": etapa,
        })
    specs.append({
        "kind": "gen",
        "field": "avance_general_vact",
        "label": "% Av. General Reforma",
    })
    return specs


def _write_legend_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.cell(1, 1, "Leyenda de estados").font = Font(bold=True, size=12, color="0F385A")
    ws.cell(2, 1, "Estado").font = Font(bold=True, color="FFFFFF")
    ws.cell(2, 2, "Significado").font = Font(bold=True, color="FFFFFF")
    for c in (1, 2):
        ws.cell(2, c).fill = _hex_fill("0F385A")
        ws.cell(2, c).alignment = Alignment(horizontal="center")
    for ri, cl in enumerate(("done", "inprog", "nostart", "devuelto", "info", "na"), 3):
        lbl = STATUS_LABEL.get(cl, cl)
        solid = STATUS_CLR.get(cl, "9aabb5")
        ws.cell(ri, 1, lbl).fill = _hex_fill(solid)
        ws.cell(ri, 1).font = Font(bold=True, color=_status_font_color(cl), size=10)
        ws.cell(ri, 1).alignment = Alignment(horizontal="center")
        ws.cell(ri, 2, lbl)
        fill6, font6 = _status_excel_style(cl)
        ws.cell(ri, 3, "Vista en tabla").fill = _hex_fill(fill6)
        ws.cell(ri, 3).font = Font(bold=True, color=font6, size=10)
        ws.cell(ri, 3).alignment = Alignment(horizontal="center")
    ws.cell(2, 3, "Vista en tabla").font = Font(bold=True, color="FFFFFF")
    ws.cell(2, 3).fill = _hex_fill("0F385A")
    ws.cell(2, 3).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16


def excel_export_bytes(df: pd.DataFrame) -> bytes:
    """Genera Excel formateado con actividades, estados y % por etapa."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if len(df) == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle VACT"
        ws.cell(1, 1, "Sin datos para exportar")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    specs = _build_export_specs(df)
    sort_cols = [c for c in ("PERIODO DE IMPLEMENTACIÓN", "NOMBRE DEL PROGRAMA") if c in df.columns]
    df_sorted = df.sort_values(sort_cols) if sort_cols else df

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle VACT"

    n_meta = sum(1 for s in specs if s["kind"] == "meta")
    col = 1

    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for spec in specs[:n_meta]:
        cell = ws.cell(1, col, spec["label"])
        cell.font = header_font
        cell.fill = _hex_fill(BRAND_PRIMARY)
        cell.alignment = header_align
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    etapa_blocks: list[tuple[str, list[dict]]] = []
    i = n_meta
    while i < len(specs):
        spec = specs[i]
        if spec["kind"] == "gen":
            break
        etapa = spec["etapa"]
        block: list[dict] = []
        while i < len(specs) and specs[i].get("etapa") == etapa and specs[i]["kind"] != "gen":
            block.append(specs[i])
            i += 1
        etapa_blocks.append((etapa, block))

    gen_spec = specs[-1] if specs and specs[-1]["kind"] == "gen" else None

    for etapa, block in etapa_blocks:
        start_col = col
        span = len(block)
        clr = _hex6(ETAPA_HEADER_CLR.get(etapa, BRAND_PRIMARY))
        short = etapa.replace(" Curricular", "")
        top = ws.cell(1, start_col, short)
        top.font = Font(bold=True, color="FFFFFF", size=10)
        top.fill = _hex_fill(clr)
        top.alignment = header_align
        if span > 1:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + span - 1)
        tints = _etapa_hdr_tints(ETAPA_HEADER_CLR.get(etapa, BRAND_PRIMARY))
        for j, bspec in enumerate(block):
            ci = start_col + j
            sub = ws.cell(2, ci, bspec["label"])
            sub.font = Font(bold=True, size=8, color=_hex6(tints["act_fg"]))
            sub.fill = _hex_fill(tints["act_bg"] if bspec["kind"] == "act" else tints["pct_bg"])
            sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col += 1

    if gen_spec:
        gcell = ws.cell(1, col, gen_spec["label"])
        gcell.font = Font(bold=True, color="FFFFFF", size=9)
        gcell.fill = _hex_fill(BRAND_PRIMARY)
        gcell.alignment = header_align
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 52
    ws.freeze_panes = "A3"

    data_start = 3
    for ri, (_, row) in enumerate(df_sorted.iterrows(), data_start):
        for ci, spec in enumerate(specs, 1):
            cell = ws.cell(ri, ci)
            cell.alignment = Alignment(
                horizontal="left" if spec["kind"] == "meta" and ci == 1 else "center",
                vertical="center",
                wrap_text=spec["kind"] == "act",
            )
            if spec["kind"] == "meta":
                cell.value = row.get(spec["field"], "—")
                cell.fill = _hex_fill(_excel_row_bg(ri, data_start))
                cell.font = Font(size=9, color="0F385A" if ci == 1 else "475569")
            elif spec["kind"] == "act":
                text, cl = _activity_display(row, spec["idx"])
                cell.value = text
                fill6, font6 = _status_excel_style(cl)
                cell.fill = _hex_fill(fill6)
                cell.font = Font(bold=True, color=font6, size=9)
            elif spec["kind"] in ("pct", "gen"):
                try:
                    pct = round(float(row.get(spec["field"], 0) or 0), 1)
                except (TypeError, ValueError):
                    pct = 0.0
                cell.value = f"{pct:.0f}%"
                fill6, font6 = _pct_excel_style(pct)
                cell.fill = _hex_fill(fill6)
                cell.font = Font(bold=True, color=font6, size=10)

    for ci in range(1, len(specs) + 1):
        letter = get_column_letter(ci)
        max_w = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, ci).value
            if v is not None:
                max_w = max(max_w, min(len(str(v)), 45))
        ws.column_dimensions[letter].width = min(max_w + 2, 42)

    ws_leg = wb.create_sheet("Leyenda")
    _write_legend_sheet(ws_leg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_heatmap_programas_etapa(df: pd.DataFrame, top_n: int = 15) -> None:
    sorted_df = df.nlargest(top_n, "avance_general_vact")
    header = (
        '<motion.div style="display:grid;grid-template-columns:180px repeat(4,1fr);gap:4px;margin-bottom:6px">'
        '<motion.div style="font-size:10px;color:#94a3b8;font-weight:700">PROGRAMA</motion.div>'
        '<motion.div style="font-size:10px;color:#94a3b8;font-weight:700;text-align:center">Alist.</motion.div>'
        '<motion.div style="font-size:10px;color:#94a3b8;font-weight:700;text-align:center">Diseño</motion.div>'
        '<motion.div style="font-size:10px;color:#94a3b8;font-weight:700;text-align:center">Desarr.</motion.div>'
        '<motion.div style="font-size:10px;color:#94a3b8;font-weight:700;text-align:center">Impl.</motion.div>'
        "</motion.div>"
    ).replace("motion.", "")
    rows = ""
    for _, row in sorted_df.iterrows():
        full = str(row.get("NOMBRE DEL PROGRAMA", "—"))
        nombre = full[:22] + (".." if len(full) > 22 else "")
        vals = [
            int(row.get(c, 0) or 0)
            for c in ("pct_alistamiento", "pct_diseno", "pct_desarrollo", "pct_implementacion")
        ]
        cells = ""
        for val in vals:
            bg = _pct_color(val)
            cells += (
                f'<motion.div style="background:{bg};opacity:0.85;height:28px;border-radius:4px;'
                f'display:flex;align-items:center;justify-content:center;font-size:10px;'
                f'font-weight:700;color:#fff">{val}%</motion.div>'
            ).replace("motion.", "")
        rows += (
            f'<motion.div style="display:grid;grid-template-columns:180px repeat(4,1fr);gap:4px;margin-bottom:4px">'
            f'<motion.div style="font-size:11px;font-weight:600;color:#0f172a;overflow:hidden;'
            f'text-overflow:ellipsis;white-space:nowrap" title="{_p_esc(full)}">{_p_esc(nombre)}</motion.div>'
            f"{cells}</motion.div>"
        ).replace("motion.", "")
    st.markdown(
        f'<motion.div style="background:#fff;border:1px solid rgba(15,56,90,.1);border-radius:12px;'
        f'padding:16px;box-shadow:0 2px 8px rgba(15,56,90,.07)">'
        f'<motion.div style="font-size:13px;font-weight:700;color:{TEXT_PRIMARY};margin-bottom:12px">'
        f"Matriz de avance — Programa × Etapa</motion.div>{header}{rows}</motion.div>".replace(
            "motion.", ""
        ),
        unsafe_allow_html=True,
    )


def render_program_ficha(df: pd.DataFrame, programa: str) -> None:
    data = get_etapas_by_programa(df, programa)
    info = data.get("info", {})
    gen = data.get("avance_general", 0)
    fac = info.get("FACULTAD_ABREV", "—")
    mod = info.get("MODALIDAD", "—")
    per = info.get("PERIODO DE IMPLEMENTACIÓN", "—")
    mod_c = MODALIDAD_CLR.get(mod, "#6e7681")
    bars = ""
    for etapa in ETAPAS_ORDEN:
        pct = data["etapas"].get(etapa, {}).get("pct", 0)
        short = etapa.replace(" Curricular", "")
        bars += (
            f'<motion.div style="margin-bottom:10px">'
            f'<motion.div style="font-size:10px;font-weight:600;color:#64748b;margin-bottom:4px">{short}</motion.div>'
            f"{p_bar_html(pct)}</motion.div>"
        ).replace("motion.", "")
    st.markdown(
        f'<motion.div style="background:#fff;border:1px solid rgba(15,56,90,.1);border-radius:12px;padding:16px">'
        f'<motion.div style="font-size:15px;font-weight:700;color:{TEXT_PRIMARY}">{_p_esc(programa)}</motion.div>'
        f'<motion.div style="margin:8px 0;display:flex;gap:8px;flex-wrap:wrap">'
        f'{badge_html(fac, info.get("FACULTAD_COLOR", "#6e7681"))}'
        f'{badge_html(mod, mod_c)}{badge_html(per, "#94a3b8")}</motion.div>'
        f'<motion.div style="font-size:24px;font-weight:800;color:{BRAND_PRIMARY}">{gen:.0f}% '
        f'<span style="font-size:12px;font-weight:600;color:{TEXT_MUTED}">avance general</span></motion.div>'
        f"{bars}</motion.div>".replace("motion.", ""),
        unsafe_allow_html=True,
    )
