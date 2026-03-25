"""
pages/6_Plan_de_Trabajo.py
Plan de trabajo sugerido con fechas de inicio y cierre por periodo propuesto.
"""

import io
import math
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from utils.data_loader import load_data, enrich_df, STATUS_LABEL

st.set_page_config(
    page_title="Plan de Trabajo · Reforma Curricular",
    page_icon="🗓️",
    layout="wide",
    initial_sidebar_state="expanded",
)

CSS = """
<style>
[data-testid="stAppViewContainer"] { background: #EEF3F8; }
[data-testid="stHeader"] {
    background: linear-gradient(135deg, #0F385A 0%, #1A5276 50%, #1FB2DE 100%) !important;
    border-bottom: 3px solid #42F2F2 !important;
}
h1,h2,h3,h4,h5  { font-family: 'Segoe UI', sans-serif; color: #0F385A !important; }
p, li, label, caption { color: #2a4a5e; }
.block-container { padding-top: 3.5rem !important; padding-bottom: 2rem; }
div[data-baseweb="select"] > div {
    background: #E3F4FB !important; border-color: rgba(31,178,222,0.45) !important;
    color: #0F385A !important; border-radius: 8px !important;
}
div[data-baseweb="select"] span    { color: #0F385A !important; }
ul[data-baseweb="menu"]            { background: #FFFFFF !important; border-radius: 8px !important; }
ul[data-baseweb="menu"] li         { color: #0F385A !important; background: #FFFFFF !important; }
ul[data-baseweb="menu"] li:hover   { background: #E3F4FB !important; }
[data-testid="stDataFrame"]        { border-radius: 10px; overflow: hidden; }
footer { visibility: hidden; }
#MainMenu { visibility: hidden; }
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F385A 0%, #1A5276 45%, #1FB2DE 100%) !important;
    border-right: none !important;
}
[data-testid="stSidebarNav"] { display: none !important; }
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label  { color: rgba(255,255,255,0.80) !important; }
[data-testid="stSidebar"] [data-testid="stPageLink"] a {
    color: rgba(255,255,255,0.82) !important; background: rgba(255,255,255,0.08) !important;
    border-radius: 8px !important; padding: 8px 12px 8px 10px !important;
    margin-bottom: 3px !important; font-size: 13px !important; font-weight: 500 !important;
}
[data-testid="stSidebar"] [data-testid="stPageLink"] a:hover {
    background: rgba(255,255,255,0.18) !important; color: #FFFFFF !important;
}
[data-testid="stSidebar"] [data-testid="stPageLink"][aria-current="page"] a,
[data-testid="stSidebar"] [data-testid="stPageLink"] a[aria-current="page"] {
    background: rgba(255,255,255,0.22) !important; color: #FFFFFF !important;
    font-weight: 700 !important; border-left: 3px solid #42F2F2 !important;
}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# ── Datos ──────────────────────────────────────────────────────────────────────
df_raw = enrich_df(load_data())

fac_abrev = {
    "Facultad de Sociedad, Cultura y Creatividad":    "FSCC",
    "Facultad de Ingeniería, Diseño e Innovación":    "FIDI",
    "Facultad de Negocios, Gestión y Sostenibilidad": "FNGS",
}

WORK_DATES = {
    "2026-2":            {"inicio": "Enero 2026",      "cierre": "Junio 2026"},
    "2027-1":            {"inicio": "Julio 2026",      "cierre": "Diciembre 2026"},
    "2027-2":            {"inicio": "Enero 2027",      "cierre": "Junio 2027"},
    "Ya está en oferta": {"inicio": "Ya implementado", "cierre": "—"},
}

PERIOD_COLORS = {
    "2026-2":            "#EC0677",
    "2027-1":            "#FBAF17",
    "2027-2":            "#A6CE38",
    "Ya está en oferta": "#1FB2DE",
}

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div style="padding:18px 6px 6px;text-align:center">'
        '<div style="font-size:16px;font-weight:700;color:#FFFFFF;line-height:1.3">'
        'Reforma Curricular</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown("<hr style='margin:10px 0'>", unsafe_allow_html=True)
    st.page_link("app.py",                              label="Resumen General",      icon="📊")
    st.page_link("pages/1_Detalle_por_Etapa.py",        label="Detalle por Etapa",    icon="📋")
    st.page_link("pages/2_Programa.py",                 label="Ficha de Programa",    icon="🔍")
    st.page_link("pages/4_Gestion_Academica.py",        label="Gestión Académica",    icon="📑")
    st.page_link("pages/5_Periodo_Propuesto.py",        label="Periodo Propuesto",    icon="📅")
    st.page_link("pages/6_Plan_de_Trabajo.py",          label="Plan de Trabajo",      icon="🗓️")
    st.markdown(
        '<div style="padding:12px 6px;font-size:10px;color:rgba(255,255,255,0.40);text-align:center">'
        'POLI · 2025–2026</div>',
        unsafe_allow_html=True,
    )

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown(
    '<div style="background:linear-gradient(135deg,#0F385A 0%,#1A5276 50%,#1FB2DE 100%);'
    'padding:18px 24px 14px;border-radius:0 0 12px 12px;border-bottom:3px solid #42F2F2;">'
    '<div style="font-size:21px;font-weight:700;color:#FFFFFF;letter-spacing:-.3px">'
    '🗓️ Plan de Trabajo Sugerido</div>'
    '<div style="font-size:12px;color:rgba(255,255,255,0.70);margin-top:5px">'
    'Fechas de inicio y cierre sugeridas según periodo de implementación propuesto</div>'
    '</div>',
    unsafe_allow_html=True,
)

st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

# ── Construir tabla completa ────────────────────────────────────────────────────
all_rows = []
for periodo, dates in WORK_DATES.items():
    sub = df_raw[df_raw["periodo_propuesto"] == periodo]
    for _, row in sub.iterrows():
        try:
            av = int(float(row.get("avance_general", 0)))
            av = 0 if math.isnan(float(av)) else av
        except Exception:
            av = 0
        cf_lbl = STATUS_LABEL.get(str(row.get("cf_st", "")), str(row.get("cf_st", "—")))
        all_rows.append({
            "Programa":             row["NOMBRE DEL PROGRAMA"],
            "Modalidad":            row.get("MODALIDAD", "—"),
            "Sede":                 row.get("SEDE", "—"),
            "Facultad":             fac_abrev.get(row.get("FACULTAD", ""), "—"),
            "Periodo propuesto":    periodo,
            "Fecha inicio sugerida": dates["inicio"],
            "Fecha cierre sugerida": dates["cierre"],
            "Avance %":             av,
            "CF":                   cf_lbl,
            "% PC":                 int(float(row.get("pc_pct", 0))),
        })

df_all = pd.DataFrame(all_rows)

# ── Filtro por Periodo ──────────────────────────────────────────────────────────
periodos_disponibles = [p for p in WORK_DATES if p in df_all["Periodo propuesto"].values]

fc1, fc2 = st.columns([2, 5])
with fc1:
    sel_periodo = st.selectbox(
        "Filtrar por periodo",
        options=["Todos"] + periodos_disponibles,
        index=0,
    )

# Aplicar filtro
if sel_periodo == "Todos":
    df_show = df_all.copy()
else:
    df_show = df_all[df_all["Periodo propuesto"] == sel_periodo].copy()

# Indicador de periodo seleccionado
color_sel = PERIOD_COLORS.get(sel_periodo, "#1FB2DE") if sel_periodo != "Todos" else "#0F385A"
dates_sel  = WORK_DATES.get(sel_periodo, {})
n_show     = len(df_show)

if sel_periodo != "Todos":
    st.markdown(
        f'<div style="background:white;border-left:5px solid {color_sel};border-radius:10px;'
        f'padding:10px 16px;margin:4px 0 10px;box-shadow:0 2px 8px rgba(15,56,90,.06);'
        f'display:flex;align-items:center;justify-content:space-between">'
        f'<div style="font-size:13px;color:#6a8a9e">'
        f'📅 Inicio sugerido: <b style="color:#0F385A">{dates_sel["inicio"]}</b>'
        f'&nbsp;·&nbsp;🏁 Cierre sugerido: <b style="color:#0F385A">{dates_sel["cierre"]}</b></div>'
        f'<div style="background:{color_sel};color:white;font-size:14px;font-weight:800;'
        f'padding:4px 14px;border-radius:16px">{n_show} programas</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
else:
    st.caption(f"{n_show} programas en total")

# ── Tabla HTML ──────────────────────────────────────────────────────────────────
def _esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _badge(text, color):
    r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    return (f'<span style="background:rgba({r},{g},{b},0.12);color:{color};font-size:10px;'
            f'font-weight:700;padding:3px 9px;border-radius:12px;white-space:nowrap">'
            f'{_esc(text)}</span>')

def _pct_bar(pct, color=None):
    pct = float(pct or 0)
    if color is None:
        color = "#5a7a2e" if pct >= 70 else ("#d97706" if pct >= 40 else "#EC0677")
    bar = "#A6CE38" if pct >= 70 else ("#FBAF17" if pct >= 40 else "#EC0677")
    return (f'<div style="min-width:64px;text-align:center">'
            f'<div style="font-size:11px;font-weight:700;color:{color};margin-bottom:2px">{int(pct)}%</div>'
            f'<div style="height:5px;background:#e2e8f0;border-radius:3px;overflow:hidden">'
            f'<div style="width:{min(pct,100):.0f}%;height:100%;background:{bar};border-radius:3px"></div>'
            f'</div></div>')

MOD_COLOR = {"Virtual": "#1FB2DE", "Presencial": "#A6CE38", "Híbrido": "#FBAF17"}
FAC_COLOR_PT = {"FSCC": "#1FB2DE", "FIDI": "#A6CE38", "FNGS": "#FBAF17"}

TH  = ('style="background:#0F385A;color:#FFFFFF;font-size:11px;font-weight:700;'
       'padding:10px 10px;text-align:center;white-space:nowrap;position:sticky;top:0;z-index:2"')
TH_L = ('style="background:#0F385A;color:#FFFFFF;font-size:11px;font-weight:700;'
        'padding:10px 14px;text-align:left;white-space:nowrap;position:sticky;top:0;z-index:2"')

rows_html = []
for idx, row in df_show.iterrows():
    row_bg = "#FFFFFF" if len(rows_html) % 2 == 0 else "#f8fafc"
    TD   = (f'style="padding:8px 10px;text-align:center;vertical-align:middle;'
            f'border-bottom:1px solid #eef3f8;background:{row_bg}"')
    TD_L = (f'style="padding:8px 14px;text-align:left;vertical-align:middle;'
            f'border-bottom:1px solid #eef3f8;background:{row_bg}"')

    mod_clr  = MOD_COLOR.get(str(row["Modalidad"]), "#9aabb5")
    fac_clr  = FAC_COLOR_PT.get(str(row["Facultad"]), "#9aabb5")
    per      = str(row["Periodo propuesto"])
    per_clr  = PERIOD_COLORS.get(per, "#9aabb5")
    cf_lbl   = _esc(str(row["CF"]))
    cf_clr   = "#5a7a2e" if "finaliz" in cf_lbl.lower() else ("#d97706" if "proceso" in cf_lbl.lower() else "#9a0050")

    rows_html.append(
        f'<tr>'
        f'<td {TD_L}><span style="font-size:12px;font-weight:600;color:#0F385A">{_esc(str(row["Programa"]))}</span></td>'
        f'<td {TD}>{_badge(str(row["Modalidad"]), mod_clr)}</td>'
        f'<td {TD}><span style="font-size:11px;color:#4a6a7e">{_esc(str(row["Sede"]))}</span></td>'
        f'<td {TD}>{_badge(str(row["Facultad"]), fac_clr)}</td>'
        f'<td {TD}>{_badge(per, per_clr)}</td>'
        f'<td {TD}><span style="font-size:11px;color:#4a6a7e">{_esc(str(row["Fecha inicio sugerida"]))}</span></td>'
        f'<td {TD}><span style="font-size:11px;color:#4a6a7e">{_esc(str(row["Fecha cierre sugerida"]))}</span></td>'
        f'<td {TD}>{_pct_bar(row["Avance %"])}</td>'
        f'<td {TD}><span style="font-size:10px;font-weight:600;color:{cf_clr}">{cf_lbl}</span></td>'
        f'<td {TD}>{_pct_bar(row["% PC"])}</td>'
        f'</tr>'
    )

table_html = (
    '<div style="overflow-x:auto;overflow-y:auto;max-height:560px;border-radius:10px;'
    'border:1px solid #dde8f0;box-shadow:0 2px 10px rgba(15,56,90,.08)">'
    '<table style="width:100%;border-collapse:collapse;font-family:\'Segoe UI\',sans-serif">'
    '<thead><tr>'
    f'<th {TH_L}>PROGRAMA</th>'
    f'<th {TH}>MODALIDAD</th>'
    f'<th {TH}>SEDE</th>'
    f'<th {TH}>FACULTAD</th>'
    f'<th {TH}>PERIODO PROPUESTO</th>'
    f'<th {TH}>FECHA INICIO</th>'
    f'<th {TH}>FECHA CIERRE</th>'
    f'<th {TH}>AVANCE %</th>'
    f'<th {TH}>CF</th>'
    f'<th {TH}>% PC</th>'
    '</tr></thead>'
    '<tbody>' + "".join(rows_html) + '</tbody>'
    '</table></div>'
)
if n_show == 0:
    st.info("Sin programas para el periodo seleccionado.")
else:
    st.markdown(table_html, unsafe_allow_html=True)

# ── Descarga Excel ──────────────────────────────────────────────────────────────
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

def _gen_excel_plan(df_exp):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de Trabajo"
    headers = list(df_exp.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="0F385A", end_color="0F385A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ri, row_data in enumerate(df_exp.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left" if ci <= 4 else "center")
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(h)) + 2, 10), 35)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

st.download_button(
    "⬇️ Descargar Plan de Trabajo (Excel)",
    data=_gen_excel_plan(df_show),
    file_name=f"plan_de_trabajo{'_' + sel_periodo if sel_periodo != 'Todos' else ''}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
