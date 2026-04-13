"""
pages/4_Gestion_Academica.py
Tabla principal de Gestión Académica con estado de todas las etapas por programa.
"""

import math
import io
import streamlit as st
import streamlit.components.v1 as _html_comp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from utils.data_loader import load_data, enrich_df, PROCESOS

st.set_page_config(
    page_title="Gestión Académica · Reforma Curricular",
    page_icon="📑",
    layout="wide",
    initial_sidebar_state="expanded",
)

CSS = """
<style>
[data-testid="stAppViewContainer"] { background: #EEF3F8; }
[data-testid="stHeader"] {
    background: linear-gradient(135deg, #0F385A 0%, #1A5276 50%, #1FB2DE 100%) !important;
    border-bottom: 3px solid #42F2F2 !important;
}
h1,h2,h3,h4,h5  { font-family: 'Segoe UI', sans-serif; color: #0F385A !important; }
p, li, label, caption { color: #2a4a5e; }
.block-container { 
    padding-top: 3.5rem !important; 
    padding-bottom: 2rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}
div[data-baseweb="select"] > div {
    background: #E3F4FB !important; border-color: rgba(31,178,222,0.45) !important;
    color: #0F385A !important; border-radius: 8px !important;
}
div[data-baseweb="select"] span    { color: #0F385A !important; }
ul[data-baseweb="menu"]            { background: #FFFFFF !important; border-radius: 8px !important; }
ul[data-baseweb="menu"] li         { color: #0F385A !important; background: #FFFFFF !important; }
ul[data-baseweb="menu"] li:hover   { background: #E3F4FB !important; }
footer { visibility: hidden; }
#MainMenu { visibility: hidden; }
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F385A 0%, #1A5276 45%, #1FB2DE 100%) !important;
    border-right: none !important;
}
[data-testid="stSidebarNav"] { display: none !important; }
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label  { color: rgba(255,255,255,0.80) !important; }
[data-testid="stSidebar"] [data-testid="stPageLink"] a {
    color: rgba(255,255,255,0.82) !important; background: rgba(255,255,255,0.08) !important;
    border-radius: 8px !important; padding: 8px 12px 8px 10px !important;
    margin-bottom: 3px !important; font-size: 13px !important; font-weight: 500 !important;
}
[data-testid="stSidebar"] [data-testid="stPageLink"] a:hover {
    background: rgba(255,255,255,0.18) !important; color: #FFFFFF !important;
}
[data-testid="stSidebar"] [data-testid="stPageLink"][aria-current="page"] a,
[data-testid="stSidebar"] [data-testid="stPageLink"] a[aria-current="page"] {
    background: rgba(255,255,255,0.22) !important; color: #FFFFFF !important;
    font-weight: 700 !important; border-left: 3px solid #42F2F2 !important;
}
[data-testid="stPills"] button {
    border: 2px solid #1A5276 !important; color: #1A5276 !important;
    background: #FFFFFF !important; border-radius: 20px !important;
    font-size: 12px !important; font-weight: 600 !important;
}
[data-testid="stPills"] button[aria-checked="true"],
[data-testid="stPills"] button[aria-pressed="true"] {
    background: #0F385A !important; color: #FFFFFF !important;
    border-color: #0F385A !important; font-weight: 700 !important;
}
/* LIMPIAR button */
[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg,#1FB2DE,#0891b2) !important;
    border-color: #1FB2DE !important;
    color: #FFFFFF !important; font-size: 11px !important; font-weight: 700 !important;
    border-radius: 8px !important; letter-spacing:.3px !important;
    box-shadow: 0 2px 8px rgba(31,178,222,0.35) !important;
}
[data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg,#0891b2,#0F385A) !important;
    border-color: #0891b2 !important;
}
[data-testid="stBaseButton-primary"] > button,
[data-testid="stBaseButton-primary"] button {
    height: 32px !important; min-height: 32px !important;
    padding: 0 10px !important; line-height: 1 !important; white-space: nowrap !important;
}
/* Align button vertically */
[data-testid="stColumn"]:has([data-testid="stBaseButton-primary"]) {
    display: flex !important; flex-direction: column !important; justify-content: center !important;
}
/* Compact filter bar */
.stVerticalBlock:has([data-testid="stPills"]) { gap: 0.1rem !important; row-gap: 0.1rem !important; }
[data-testid="stPills"] { padding-bottom: 0 !important; margin-bottom: 0 !important; min-height: unset !important; }
[data-testid="stHorizontalBlock"]:has([data-testid="stPills"]) { align-items: center !important; padding-bottom: 0 !important; margin-bottom: 0 !important; }
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# ── Datos ──────────────────────────────────────────────────────────────────────
df_raw = enrich_df(load_data())

# Filtro de Nivel homologado
niveles = [n for n in ["Pregrado", "Posgrado"] if "NIVEL_HOMOLOGADO" in df_raw.columns and n in df_raw["NIVEL_HOMOLOGADO"].values]

fac_abrev = {
    "Facultad de Sociedad, Cultura y Creatividad":    "FSCC",
    "Facultad de Ingeniería, Diseño e Innovación":    "FIDI",
    "Facultad de Negocios, Gestión y Sostenibilidad": "FNGS",
}
FAC_COLOR = {"FSCC": "#1FB2DE", "FIDI": "#A6CE38", "FNGS": "#FBAF17"}
MOD_COLOR = {"Virtual": "#1FB2DE", "Presencial": "#A6CE38", "Híbrido": "#FBAF17"}
PER_COLOR = {
    "2026-2":            "#EC0677",
    "2027-1":            "#FBAF17",
    "2027-2":            "#5a7a2e",
    "Ya está en oferta": "#1FB2DE",
}
PER_DISPLAY = {
    "Ya está en oferta": "Ya oferta",
    "2026-2": "2026-2",
    "2027-1": "2027-1",
    "2027-2": "2027-2",
}
PER_REVERSE = {v: k for k, v in PER_DISPLAY.items()}

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div style="padding:18px 6px 6px;text-align:center">'
        '<div style="font-size:16px;font-weight:700;color:#FFFFFF;line-height:1.3">'
        'Reforma Curricular</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown("<hr style='margin:10px 0'>", unsafe_allow_html=True)
    st.page_link("app.py",                              label="Resumen General",      icon="📊")
    st.page_link("pages/1_Detalle_por_Etapa.py",        label="Detalle por Etapa",    icon="📋")
    st.page_link("pages/2_Programa.py",                 label="Resumen Programa",     icon="🔍")
    st.markdown(
        '<div style="padding:12px 6px;font-size:10px;color:rgba(255,255,255,0.40);text-align:center">'
        'POLI · 2025–2026</div>',
        unsafe_allow_html=True,
    )

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown(
    '<div style="background:linear-gradient(135deg,#0F385A 0%,#1A5276 50%,#1FB2DE 100%);'
    'padding:18px 24px 14px;border-radius:0 0 12px 12px;border-bottom:3px solid #42F2F2;">'
    '<div style="font-size:21px;font-weight:700;color:#FFFFFF;letter-spacing:-.3px">'
    '📑 Gestión Académica</div>'
    '<div style="font-size:12px;color:rgba(255,255,255,0.70);margin-top:5px">'
    'Estado de etapas por programa · Avance general calculado</div>'
    '</div>',
    unsafe_allow_html=True,
)
st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# ── Filtros ────────────────────────────────────────────────────────────────────
def _clear():
    st.session_state["ga_buscar"] = ""
    st.session_state["ga_mod"]    = []
    st.session_state["ga_per"]    = []
    st.session_state["ga_nivel"]  = []

_use_pills = hasattr(st, "pills")
_per_opts  = list(PER_DISPLAY.values())

_LBL = (
    'style="padding-top:8px;font-size:11px;font-weight:700;color:#0F385A;'
    'letter-spacing:.4px;white-space:nowrap"'
)

with st.container():
    # ── Fila 1: BUSCAR · MODALIDAD · LIMPIAR ───────────────────────────────────
    lb1, in1, _sp, lb2, in2, btn_col = st.columns([0.55, 2.6, 0.05, 0.7, 2.75, 0.65])
    with lb1:
        st.markdown(f'<div {_LBL}>🔍 BUSCAR</div>', unsafe_allow_html=True)
    with in1:
        st.text_input("buscar", placeholder="Programa o escuela...",
                      key="ga_buscar", label_visibility="collapsed")
    with lb2:
        st.markdown(f'<div {_LBL}>📋 MODALIDAD</div>', unsafe_allow_html=True)
    with in2:
        if _use_pills:
            st.pills("mod", ["Virtual", "Presencial", "Híbrido"],
                     selection_mode="multi", key="ga_mod", label_visibility="collapsed")
        else:
            st.multiselect("mod", ["Virtual", "Presencial", "Híbrido"],
                           key="ga_mod", label_visibility="collapsed", placeholder="Todas")
    with btn_col:
        st.button("✕ LIMPIAR", on_click=_clear, type="primary")

    # ── Fila 2: PERÍODO · NIVEL · Mostrando ────────────────────────────────────────────
    lb3, in3, lb4, in4, cnt_col = st.columns([0.55, 2.6, 0.7, 2.5, 2.15])
    with lb3:
        st.markdown(f'<div {_LBL}>📅 PERÍODO</div>', unsafe_allow_html=True)
    with in3:
        if _use_pills:
            st.pills("per", _per_opts,
                     selection_mode="multi", key="ga_per", label_visibility="collapsed")
        else:
            st.multiselect("per", _per_opts,
                           key="ga_per", label_visibility="collapsed", placeholder="Todos")
    with lb4:
        st.markdown(f'<div {_LBL}>🎓 NIVEL</div>', unsafe_allow_html=True)
    with in4:
        if _use_pills:
            st.pills("nivel", niveles, selection_mode="multi", key="ga_nivel", label_visibility="collapsed")
        else:
            st.multiselect("nivel", niveles, key="ga_nivel", label_visibility="collapsed", placeholder="Todos")
    with cnt_col:
        _counter = st.empty()

# ── Aplicar filtros ────────────────────────────────────────────────────────────
df = df_raw.copy()

buscar_v  = (st.session_state.get("ga_buscar") or "").strip().lower()
sel_mod_v = list(st.session_state.get("ga_mod") or [])
sel_per_v = [PER_REVERSE.get(p, p) for p in (st.session_state.get("ga_per") or [])]
sel_nivel_v = list(st.session_state.get("ga_nivel") or [])

if buscar_v:
    _fal = {k.lower(): v for k, v in fac_abrev.items()}
    mask_prog = df["NOMBRE DEL PROGRAMA"].str.lower().str.contains(buscar_v, na=False)
    mask_fac  = df["FACULTAD"].str.lower().str.contains(buscar_v, na=False)
    mask_fabr = df["FACULTAD"].str.lower().map(_fal).str.lower().str.contains(buscar_v, na=False)
    df = df[mask_prog | mask_fac | mask_fabr]

if sel_mod_v:
    df = df[df["MODALIDAD"].isin(sel_mod_v)]
if sel_per_v:
    df = df[df["periodo_propuesto"].isin(sel_per_v)]
if sel_nivel_v:
    df = df[df["NIVEL_HOMOLOGADO"].isin(sel_nivel_v)]

n_total = len(df_raw)
n_show  = len(df)

_counter.markdown(
    f'<div style="padding-top:9px;font-size:12px;color:#6a8a9e;text-align:right;white-space:nowrap">'
    f'Mostrando <b style="color:#0F385A">{n_show}</b> de '
    f'<b style="color:#0F385A">{n_total}</b></div>',
    unsafe_allow_html=True,
)

# ── Helpers HTML ───────────────────────────────────────────────────────────────
def _esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _proc_icon(val):
    if val is None or (isinstance(val, float) and math.isnan(float(val))):
        return '<span style="color:#9aabb5;font-size:16px">—</span>'
    v = float(val)
    if v >= 100:
        return '<span style="font-size:16px">✅</span>'
    if v > 0:
        return '<span style="font-size:16px">⚠️</span>'
    return '<span style="font-size:16px">🔴</span>'

def _syl_icon(syl):
    if syl == "Si":
        return '<span style="font-size:16px">✅</span>'
    if syl == "NO":
        return '<span style="font-size:16px">🔴</span>'
    return '<span style="color:#9aabb5;font-size:16px">—</span>'

def _pct_bar(pct):
    try:
        pct = float(pct if pct is not None else 0)
    except Exception:
        pct = 0.0
    import math as _m
    if _m.isnan(pct): pct = 0.0
    clr = "#5a7a2e" if pct >= 70 else ("#d97706" if pct >= 40 else "#EC0677")
    bg  = "#f0f8e8" if pct >= 70 else ("#fef9e8" if pct >= 40 else "#fce8f2")
    bar_clr = "#A6CE38" if pct >= 70 else ("#FBAF17" if pct >= 40 else "#EC0677")
    return (
        f'<div style="min-width:64px;text-align:center">'
        f'<div style="font-size:11px;font-weight:700;color:{clr};margin-bottom:3px">{int(pct)}%</div>'
        f'<div style="height:5px;background:#e2e8f0;border-radius:3px;overflow:hidden">'
        f'<div style="width:{min(pct,100):.0f}%;height:100%;background:{bar_clr};border-radius:3px"></div>'
        f'</div></div>'
    )

def _badge(text, color):
    r = int(color[1:3], 16); g = int(color[3:5], 16); b = int(color[5:7], 16)
    bg = f"rgba({r},{g},{b},0.12)"
    return (f'<span style="background:{bg};color:{color};font-size:10px;font-weight:700;'
            f'padding:3px 9px;border-radius:12px;white-space:nowrap">{_esc(text)}</span>')

# ── Construir tabla HTML ───────────────────────────────────────────────────────
TH = ('style="background:#0F385A;color:#FFFFFF;font-size:11px;font-weight:700;'
      'padding:10px 8px;text-align:center;white-space:nowrap;position:sticky;top:0;z-index:2;'
      'border-right:1px solid rgba(255,255,255,0.10)"')
TH_L = ('style="background:#0F385A;color:#FFFFFF;font-size:11px;font-weight:700;'
        'padding:10px 14px;text-align:left;white-space:nowrap;position:sticky;top:0;z-index:2;'
        'border-right:1px solid rgba(255,255,255,0.10)"')

ETAPA_COLS = [
    ("G.ACADÉMICA",  f"proc_Gestión Académica",                      "proc"),
    ("C.FINANCIERO", f"proc_Gestión Financiera",                      "proc"),
    ("ASEGURAMIENT", f"proc_Aseguramiento de la Calidad",             "proc"),
    ("SYLLABUS",     "syl_val",                                       "syl"),
    ("PROD.CONT.",   "pc_pct",                                        "bar"),
    ("CONVENIOS",    f"proc_Convenios Institucionales",               "proc"),
    ("BANNER",       "ban_pct",                                       "bar"),
    ("TIPO DE TRÁMITE", "Tipo de trámite de aseguramiento de la calidad", "text"),
    ("FECHA NOTIF. MEN", "Fecha de\nDocumentos de notificación MEN", "text"),
    ("REQ. MINISTERIAL", "¿Requiere aprobación ministerial?",      "text"),
]

rows_html = []
for idx, (_, row) in enumerate(df.iterrows()):
    row_bg = "#FFFFFF" if idx % 2 == 0 else "#f8fafc"
    TD  = (f'style="padding:8px;text-align:center;vertical-align:middle;'
           f'border-bottom:1px solid #eef3f8;background:{row_bg}"')
    TD_L = (f'style="padding:8px 14px;text-align:left;vertical-align:middle;'
            f'border-bottom:1px solid #eef3f8;background:{row_bg};min-width:200px;max-width:280px"')

    prog    = _esc(row.get("NOMBRE DEL PROGRAMA", "—"))
    fac     = fac_abrev.get(str(row.get("FACULTAD", "")), "—")
    fac_clr = FAC_COLOR.get(fac, "#9aabb5")
    mod     = str(row.get("MODALIDAD", "—"))
    mod_clr = MOD_COLOR.get(mod, "#9aabb5")
    sede    = _esc(str(row.get("SEDE", "—")))
    per     = str(row.get("periodo_propuesto", "—"))
    per_clr = PER_COLOR.get(per, "#9aabb5")
    per_lbl = PER_DISPLAY.get(per, per)

    prog_cell = (
        f'<div style="font-size:12px;font-weight:600;color:#0F385A;line-height:1.4">{prog}</div>'
        f'<div style="font-size:10px;color:{fac_clr};font-weight:600;margin-top:2px">{fac}</div>'
    )

    etapa_cells = []
    for _, col_key, typ in ETAPA_COLS:
        val = row.get(col_key)
        if typ == "proc":
            try:
                v = None if (val == "" or val is None) else float(val)
                if v is not None and math.isnan(v):
                    v = None
            except Exception:
                v = None
            etapa_cells.append(f'<td {TD}>{_proc_icon(v)}</td>')
        elif typ == "syl":
            etapa_cells.append(f'<td {TD}>{_syl_icon(str(val or "N/A"))}</td>')
        elif typ == "bar":
            try:
                pct = float(val if val is not None else 0)
                etapa_cells.append(f'<td {TD}>{_pct_bar(pct)}</td>')
            except Exception:
                etapa_cells.append(f'<td {TD}><span style="color:#9aabb5;font-size:16px">—</span></td>')
        else:
            # Para columnas de texto (aseguramiento de la calidad)
            etapa_cells.append(f'<td {TD}><span style="font-size:11px;color:#0F385A">{val if val not in [None, "", "nan"] else "—"}</span></td>')

    rows_html.append(
        f'<tr>'
        f'<td {TD_L}>{prog_cell}</td>'
        f'<td {TD}>{_badge(mod, mod_clr)}</td>'
        f'<td {TD}><span style="font-size:11px;color:#4a6a7e">{sede}</span></td>'
        f'<td {TD}>{_badge(per_lbl, per_clr)}</td>'
        + "".join(etapa_cells) +
        f'</tr>'
    )

if n_show == 0:
    st.info("Sin programas para los filtros seleccionados.")
else:
    header_cells = "".join(f'<th {TH}>{h}</th>' for h, _, _ in ETAPA_COLS)
    table_html = (
        '<div style="overflow-x:auto;overflow-y:auto;max-height:580px;border-radius:12px;'
        'border:1.5px solid #b5c9d8;box-shadow:0 2px 12px rgba(15,56,90,.10);background:#fafdff">'
        '<table style="width:100%;border-collapse:separate;border-spacing:0;font-family:\'Segoe UI\',sans-serif">'
        '<thead><tr>'
        f'<th {TH_L}>PROGRAMA</th>'
        f'<th {TH}>MODALIDAD</th>'
        f'<th {TH}>SEDE</th>'
        f'<th {TH}>PERÍODO</th>'
        + header_cells +
        '</tr></thead>'
        '<tbody>'
        + "".join(rows_html) +
        '</tbody></table></div>'
    )
    _tph = max(300, min(1400, 100 + n_show * 60))
    _html_comp.html(table_html, height=_tph, scrolling=True)

# ── Descarga Excel ─────────────────────────────────────────────────────────────
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

def _gen_excel(df_src):
    rows_exp = []
    for _, row in df_src.iterrows():
        fac = fac_abrev.get(str(row.get("FACULTAD", "")), "—")
        r   = {
            "Programa":  row.get("NOMBRE DEL PROGRAMA", "—"),
            "Facultad":  fac,
            "Modalidad": row.get("MODALIDAD", "—"),
            "Sede":      row.get("SEDE", "—"),
            "Período":   PER_DISPLAY.get(str(row.get("periodo_propuesto", "—")), "—"),
            "Avance %":  int(float(row.get("avance_general", 0) or 0)),
        }
        for hdr, col_key, typ in ETAPA_COLS:
            val = row.get(col_key)
            if typ == "proc":
                try:
                    v = float(val)
                    if math.isnan(v):
                        r[hdr] = "N/A"
                    elif v >= 100:
                        r[hdr] = "Finalizado"
                    elif v > 0:
                        r[hdr] = f"En proceso ({int(v)}%)"
                    else:
                        r[hdr] = "Sin iniciar"
                except Exception:
                    r[hdr] = "N/A"
            elif typ == "syl":
                r[hdr] = str(val or "N/A")
            else:
                try:
                    r[hdr] = f"{int(float(val or 0))}%"
                except Exception:
                    r[hdr] = "0%"
        rows_exp.append(r)

    df_exp = pd.DataFrame(rows_exp)
    wb = Workbook(); ws = wb.active; ws.title = "Gestión Académica"
    headers = list(df_exp.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="0F385A", end_color="0F385A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ri, row_data in enumerate(df_exp.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left" if ci <= 5 else "center")
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(h)) + 2, 8), 35)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

st.download_button(
    "⬇️ Descargar Excel",
    data=_gen_excel(df),
    file_name="gestion_academica.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

