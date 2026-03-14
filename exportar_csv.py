"""
exportar_csv.py
Exporta la hoja Borrador del Excel a data/raw/control_maestro.csv.
Ejecutar cuando el archivo Excel se actualice:
    python3 exportar_csv.py
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Instala openpyxl:  pip install openpyxl")

EXCEL_PATH = Path("CONTROL_MAESTRO_DE_REFORMA_CURRICULAR.xlsx")
OUT_PATH   = Path("data/raw/control_maestro.csv")
SHEET_NAME = "Borrador"
HEADER_ROW = 10   # fila 10 (índice 0-based = 9) contiene los encabezados
DATA_START = 11   # datos desde fila 11
ID_COL     = 2    # columna FACULTAD — si está vacía, la fila se omite


def export():
    if not EXCEL_PATH.exists():
        sys.exit(f"No se encuentra el archivo: {EXCEL_PATH}")

    print(f"Abriendo {EXCEL_PATH} …")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"No se encontró la hoja '{SHEET_NAME}'. Hojas disponibles: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # Leer encabezados (fila 10)
    headers_raw = [cell.value for cell in ws[HEADER_ROW]]
    headers = [
        str(h).strip().replace("\n", " ") if h else f"col_{i}"
        for i, h in enumerate(headers_raw)
    ]

    # Leer filas de datos
    rows = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        if row[ID_COL] is None:
            continue
        clean = []
        for v in row:
            if v is None:
                clean.append("")
            elif hasattr(v, "strftime"):
                clean.append(v.strftime("%Y-%m-%d"))
            else:
                clean.append(str(v).strip())
        rows.append(clean)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f"CSV guardado en {OUT_PATH}")
    print(f"  Filas: {len(rows)}")
    print(f"  Columnas: {len(headers)}")


if __name__ == "__main__":
    export()
