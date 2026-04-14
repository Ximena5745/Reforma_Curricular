"""
app.py  —  Control Maestro de Reforma Curricular
Página principal: Resumen General
"""

import streamlit as st
import streamlit.components.v1 as _html_comp
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import math
from utils.data_loader import (
    load_data, enrich_df, apply_filters, ETAPAS_MAP, PROCESOS,
    PROCESO_COLOR, STATUS_LABEL, STATUS_COLOR, color_for_pct,
)

st.set_page_config(
    page_title="Reforma Curricular · POLI",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS modo claro · colores institucionales ───────────────────────────────────
st.markdown("""
<style>
/* ── Fondos generales ── */
[data-testid="stAppViewContainer"] { background: #EEF3F8; }
/* Header de Streamlit (barra superior delgada) coincide con el degradado */
[data-testid="stHeader"] {
    background: linear-gradient(135deg, #0F385A 0%, #1A5276 50%, #1FB2DE 100%) !important;
    border-bottom: 3px solid #42F2F2 !important;
}
h1,h2,h3,h4,h5                     { font-family: 'Segoe UI', sans-serif; color: #0F385A !important; }
p, li, label, caption               { color: #2a4a5e; }
[data-testid="stCaption"]           { color: #6a8a9e !important; }
/* ── Block container ── */
.block-container { padding-top: 3.5rem !important; padding-bottom: 2rem; }
/* ── Selectbox — fondo diferenciado azul muy claro ── */
div[data-baseweb="select"] > div {
    background: #E3F4FB !important;
    border-color: rgba(31,178,222,0.45) !important;
    color: #0F385A !important;
    border-radius: 8px !important;
}
div[data-baseweb="select"] span    { color: #0F385A !important; }
[data-testid="stSelectbox"] label  { font-size: 12px; color: #4a6a7e; }
/* ── Dropdown list (options) ── */
ul[data-baseweb="menu"]            { background: #FFFFFF !important; border: 1px solid rgba(31,178,222,0.30) !important; box-shadow: 0 6px 20px rgba(15,56,90,0.14) !important; border-radius: 8px !important; }
ul[data-baseweb="menu"] li         { color: #0F385A !important; background: #FFFFFF !important; }
ul[data-baseweb="menu"] li:hover   { background: #E3F4FB !important; }
li[aria-selected="true"]           { background: #d0ecf8 !important; font-weight: 600 !important; }
/* ── Tabs ── */
button[data-baseweb="tab"]         { color: #6a8a9e !important; background: transparent !important; font-size: 13px !important; font-weight: 500 !important; }
button[data-baseweb="tab"][aria-selected="true"] { color: #0F385A !important; border-bottom-color: #1FB2DE !important; font-weight: 700 !important; }
[data-testid="stDataFrame"]        { border-radius: 10px; overflow: hidden; box-shadow: 0 1px 6px rgba(15,56,90,0.08); }
hr                                  { border-color: rgba(15,56,90,0.10) !important; margin: 6px 0 !important; }
footer { visibility: hidden; }
#MainMenu { visibility: hidden; }
[data-testid="stExpander"]         { background: #FFFFFF; border: 1px solid rgba(15,56,90,0.10); border-radius: 10px; }
/* Info box */
[data-testid="stNotification"]     { background: #e8f6fc !important; color: #0F385A !important; border-color: #1FB2DE !important; }
/* ── Sidebar con degradado ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F385A 0%, #1A5276 45%, #1FB2DE 100%) !important;
    border-right: none !important;
}
[data-testid="stSidebarNav"] { display: none !important; }
/* Sidebar text */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label  { color: rgba(255,255,255,0.80) !important; }
[data-testid="stSidebar"] hr     { border-color: rgba(255,255,255,0.20) !important; }
/* Page link buttons en sidebar */
[data-testid="stSidebar"] [data-testid="stPageLink"] a {
    color: rgba(255,255,255,0.82) !important;
    background: rgba(255,255,255,0.08) !important;
    border-radius: 8px !important;
    padding: 8px 12px 8px 10px !important;
    margin-bottom: 3px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    transition: background .15s;
}
[data-testid="stSidebar"] [data-testid="stPageLink"] a:hover {
    background: rgba(255,255,255,0.18) !important;
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] [data-testid="stPageLink"][aria-current="page"] a,
[data-testid="stSidebar"] [data-testid="stPageLink"] a[aria-current="page"] {
    background: rgba(255,255,255,0.22) !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    border-left: 3px solid #42F2F2 !important;
}
/* Download button */
[data-testid="stDownloadButton"] > button {
    background: #1FB2DE !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #0F385A !important;
}
/* ── Pills / botones de filtro ── */
[data-testid="stPills"] button,
[data-testid="stSegmentedControl"] button {
    border: 2px solid #1A5276 !important;
    color: #1A5276 !important;
    background: #FFFFFF !important;
    border-radius: 20px !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    transition: all .15s !important;
    box-shadow: 0 1px 3px rgba(15,56,90,0.12) !important;
}
[data-testid="stPills"] button:hover,
[data-testid="stSegmentedControl"] button:hover {
    border-color: #1FB2DE !important;
    background: #1FB2DE !important;
    color: #FFFFFF !important;
}
[data-testid="stPills"] button[aria-checked="true"],
[data-testid="stPills"] button[aria-pressed="true"],
[data-testid="stPills"] button[data-active="true"],
[data-testid="stSegmentedControl"] button[aria-selected="true"] {
    background: #0F385A !important;
    color: #FFFFFF !important;
    border-color: #0F385A !important;
    font-weight: 700 !important;
    box-shadow: 0 2px 6px rgba(15,56,90,0.30) !important;
}
/* LIMPIAR button */
[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg,#1FB2DE,#0891b2) !important;
    border-color: #1FB2DE !important;
    color: #FFFFFF !important; font-size: 11px !important; font-weight: 700 !important;
    border-radius: 8px !important; letter-spacing:.3px !important;
    box-shadow: 0 2px 8px rgba(31,178,222,0.35) !important;
}
[data-testid="stBaseButton-primary"] > button,
[data-testid="stBaseButton-primary"] button {
    height: 32px !important; min-height: 32px !important;
    padding: 0 10px !important; line-height: 1 !important; white-space: nowrap !important;
}
/* Align button vertically */
[data-testid="stColumn"]:has([data-testid="stBaseButton-primary"]) {
    display: flex !important; flex-direction: column !important; justify-content: center !important;
}
[data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg,#0891b2,#0F385A) !important;
    border-color: #0891b2 !important;
}
/* ── Compact filter bar: collapse gap between pill rows ── */
.stVerticalBlock:has([data-testid="stPills"]) {
    gap: 0.1rem !important;
    row-gap: 0.1rem !important;
}
[data-testid="stPills"] {
    padding-bottom: 0 !important;
    margin-bottom: 0 !important;
    min-height: unset !important;
}
[data-testid="stHorizontalBlock"]:has([data-testid="stPills"]) {
    align-items: center !important;
    padding-bottom: 0 !important;
    margin-bottom: 0 !important;
}
/* Remove excess space below filter block and before tabs */
.stVerticalBlock:has([data-testid="stPills"]) + .stVerticalBlock {
    margin-top: 0 !important;
    padding-top: 0 !important;
}
.stMainBlockContainer [data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"]:has([data-testid="stPills"]) {
    margin-bottom: 2px !important;
    padding-bottom: 2px !important;
}
/* Tighten gap between top-level page blocks (header → filters → tabs) */
.stMainBlockContainer > div > div > .stVerticalBlock { gap: 0.4rem !important; }
/* Remove top margin on tabs bar */
[data-testid="stTabs"] { margin-top: 0 !important; padding-top: 0 !important; }
[data-baseweb="tab-list"] { margin-top: 0 !important; }
/* Sidebar toggle always visible */
[data-testid="stSidebarCollapsedControl"] {
    display: flex !important; visibility: visible !important;
    opacity: 1 !important; position: fixed !important;
    top: 0.5rem !important; left: 0.5rem !important; z-index: 999999 !important;
}
/* Remove white divider between header elements */
[data-testid="stHorizontalBlock"]:has([data-testid="stPills"]) + div {
    margin-top: 0 !important;
}
/* Pill colors - Modalidad (nth-child per option) */
[data-testid="stPills"] button[aria-checked="false"]:nth-child(1),
[data-testid="stPills"] button[aria-pressed="false"]:nth-child(1) {
    border-color: #FBAF17 !important; color: #d97706 !important;
}
[data-testid="stPills"] button[aria-checked="false"]:nth-child(2),
[data-testid="stPills"] button[aria-pressed="false"]:nth-child(2) {
    border-color: #A6CE38 !important; color: #5a7a2e !important;
}
[data-testid="stPills"] button[aria-checked="false"]:nth-child(3),
[data-testid="stPills"] button[aria-pressed="false"]:nth-child(3) {
    border-color: #1FB2DE !important; color: #0891b2 !important;
}
[data-testid="stPills"] button[aria-checked="false"]:nth-child(4),
[data-testid="stPills"] button[aria-pressed="false"]:nth-child(4) {
    border-color: #7c3aed !important; color: #7c3aed !important;
}
/* Compact filter - reduce height */
[data-testid="stHorizontalBlock"]:has([data-testid="stPills"]) {
    min-height: unset !important; height: auto !important;
    padding-top: 0 !important; padding-bottom: 0 !important;
}
</style>
""", unsafe_allow_html=True)

# ── Datos ──────────────────────────────────────────────────────────────────────
df_raw = enrich_df(load_data())

fac_labels = {
    "Facultad de Sociedad, Cultura y Creatividad":    "Sociedad, Cultura y Creatividad",
    "Facultad de Ingeniería, Diseño e Innovación":    "Ingeniería, Diseño e Innovación",
    "Facultad de Negocios, Gestión y Sostenibilidad": "Negocios, Gestión y Sostenibilidad",
}
fac_abrev = {
    "Facultad de Sociedad, Cultura y Creatividad":    "FSCC",
    "Facultad de Ingeniería, Diseño e Innovación":    "FIDI",
    "Facultad de Negocios, Gestión y Sostenibilidad": "FNGS",
}
fac_inv       = {v: k for k, v in fac_labels.items()}
fac_abrev_inv = {v: k for k, v in fac_abrev.items()}
FAC_LIST  = list(fac_labels.values())
FAC_PALETTE = {
    "Sociedad, Cultura y Creatividad":   "#EC0677",
    "Ingeniería, Diseño e Innovación":   "#1FB2DE",
    "Negocios, Gestión y Sostenibilidad":"#A6CE38",
}

# ── Sidebar personalizado ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div style="padding:18px 6px 6px;text-align:center">'
        '<div style="font-size:16px;font-weight:700;color:#FFFFFF;line-height:1.3">'
        'Reforma Curricular</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown("<hr style='margin:10px 0'>", unsafe_allow_html=True)
    st.page_link("app.py",                              label="Resumen General",      icon="📊")
    st.page_link("pages/1_Detalle_por_Etapa.py",        label="Detalle por Etapa",    icon="📋")
    st.page_link("pages/2_Programa.py",                 label="Resumen Programa",     icon="🔍")
    st.markdown("<hr style='margin:10px 0'>", unsafe_allow_html=True)
    st.markdown("<div style='flex:1'></div>", unsafe_allow_html=True)
    st.markdown(
        '<div style="padding:12px 6px;font-size:10px;color:rgba(255,255,255,0.40);text-align:center">'
        'POLI · 2025–2026</div>',
        unsafe_allow_html=True,
    )

# ── Header banner ───────────────────────────────────────────────────────────────
# El stHeader (barra de Streamlit) ya tiene el degradado y cubre todo el ancho.
# Este banner es la continuación dentro del contenido principal.
st.markdown(
    '<div style="'
    'background:linear-gradient(135deg,#0F385A 0%,#1A5276 50%,#1FB2DE 100%);'
    'padding:18px 24px 14px;'
    'margin:0 0 0 0;'
    'border-radius:0 0 12px 12px;'
    'border-bottom:3px solid #42F2F2;">'
    '<div style="font-size:21px;font-weight:700;color:#FFFFFF;letter-spacing:-.3px">'
    'Reforma Curricular de Programas Académicos Poli</div>'
    '<div style="font-size:12px;color:rgba(255,255,255,0.70);margin-top:5px">'
    'Seguimiento de avance por proceso y etapa</div>'
    '</div>',
    unsafe_allow_html=True,
)

# ── Opciones y helpers de filtro (reutilizados en tabs) ─────────────────────────
_use_pills = hasattr(st, "pills")
_mods_ops  = sorted(df_raw["MODALIDAD"].dropna().unique().tolist())
fac_ops    = sorted([fac_abrev.get(f, f) for f in df_raw["FACULTAD"].dropna().unique()])
_pers_ops  = sorted(df_raw["PERIODO DE IMPLEMENTACIÓN"].dropna().unique().tolist())
_niveles_ops = [n for n in ["Pregrado", "Posgrado"] if n in df_raw.get("NIVEL_HOMOLOGADO", pd.Series(dtype=str)).values]

def _clear_app():
    st.session_state["flt_mod"]   = []
    st.session_state["flt_fac"]   = []
    st.session_state["flt_per"]   = []
    st.session_state["flt_nivel"] = []

_LBL = ('style="padding-top:8px;font-size:11px;font-weight:700;color:#0F385A;'
        'letter-spacing:.4px;white-space:nowrap"')

# ── Filtrar datos usando session_state (widgets se renderizan dentro de cada tab) ──
_sel_mod   = list(st.session_state.get("flt_mod")   or [])
_sel_fac   = list(st.session_state.get("flt_fac")   or [])
_sel_per   = list(st.session_state.get("flt_per")   or [])
_sel_nivel = list(st.session_state.get("flt_nivel") or [])
modalidad_f = _sel_mod
facultad_f  = [fac_abrev_inv.get(f, f) for f in _sel_fac]
periodo_f   = _sel_per
df = apply_filters(df_raw.copy(), modalidad_f, facultad_f, periodo_f)
if _sel_nivel and "NIVEL_HOMOLOGADO" in df.columns:
    df = df[df["NIVEL_HOMOLOGADO"].isin(_sel_nivel)]
n  = len(df)

# ── Cálculos previos (no rendering) ────────────────────────────────────────────
all_cl   = []
for i in range(len(ETAPAS_MAP)):
    col = f"cl_{i}"
    if col in df.columns:
        all_cl.extend(df[col].tolist())
avg_av   = int(df["avance_general"].mean()) if n > 0 else 0
cnt_adv  = int((df["avance_general"] >= 70).sum())

def clasificar_programa(avance, periodo):
    """Clasificación combinada A+C por avance y periodo de implementación."""
    periodo = str(periodo) if pd.notna(periodo) else ""
    es_2026_o_oferta = "2026" in periodo or "oferta" in periodo.lower() or "Ya en oferta" in periodo
    es_2027_1 = "2027-1" in periodo
    if es_2026_o_oferta and avance < 70:
        return "Urgente"
    if es_2027_1 and avance < 40:
        return "Prioritario"
    if avance < 70:
        return "En seguimiento"
    return "En curso"

if n > 0:
    df["_clasif"] = df.apply(
        lambda r: clasificar_programa(r["avance_general"], r["PERIODO DE IMPLEMENTACIÓN"]), axis=1
    )
    cnt_urgente    = int((df["_clasif"] == "Urgente").sum())
    cnt_prioritario = int((df["_clasif"] == "Prioritario").sum())
    cnt_crit       = cnt_urgente + cnt_prioritario   # para compatibilidad
else:
    df["_clasif"] = "En curso"
    cnt_urgente = cnt_prioritario = cnt_crit = 0
cnt_2026   = int(df["PERIODO DE IMPLEMENTACIÓN"].str.contains("2026",   na=False).sum())
cnt_2027_1 = int(df["PERIODO DE IMPLEMENTACIÓN"].str.contains("2027-1", na=False).sum())
proc_avgs  = {p: (df[f"proc_{p}"].dropna().mean() if df[f"proc_{p}"].dropna().shape[0] > 0 else 0)
              for p in PROCESOS}
proc_min     = min(proc_avgs, key=proc_avgs.get)
proc_max     = max(proc_avgs, key=proc_avgs.get)
proc_min_pct = int(proc_avgs[proc_min])
proc_max_pct = int(proc_avgs[proc_max])

proc_short_map = {
    "Gestión Académica": "Gest. Académica",
    "Gestión Financiera": "Gest. Financiera",
    "Aseguramiento de la Calidad": "Aseg. Calidad",
    "Producción de Contenidos": "Prod. Contenidos",
    "Convenios Institucionales": "Convenios",
    "Parametrizar Reforma en Banner": "Banner",
    "Publicación en Página Web": "Pág. Web",
}

pct_avgs, done_l, inp_l, nst_l, na_l = [], [], [], [], []
for proc in PROCESOS:
    p_vals = df[f"proc_{proc}"]
    pct_avgs.append(int(p_vals.dropna().mean()) if p_vals.dropna().any() else 0)
    done_l.append(int((p_vals == 100).sum()))
    inp_l .append(int(((p_vals > 0) & (p_vals < 100)).sum()))
    nst_l .append(int((p_vals == 0).sum()))
    na_l  .append(int(p_vals.isna().sum()))

# ── Helper functions ────────────────────────────────────────────────────────────
def _arc(pct, color, r=22, sz=56):
    circ = 2 * math.pi * r
    dash = circ * min(pct, 100) / 100
    gap  = circ - dash
    c    = sz // 2
    return (
        f'<svg width="{sz}" height="{sz}" viewBox="0 0 {sz} {sz}">'
        f'<circle cx="{c}" cy="{c}" r="{r}" fill="none" stroke="rgba(15,56,90,0.10)" stroke-width="5"/>'
        f'<circle cx="{c}" cy="{c}" r="{r}" fill="none" stroke="{color}" stroke-width="5"'
        f' stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-linecap="round"'
        f' transform="rotate(-90 {c} {c})"/>'
        f'</svg>'
    )

def _kpi(label, val, sub, color, pct=None, icon="◈", tooltip=""):
    arc = _arc(pct, color) if pct is not None else (
        f'<div style="width:56px;height:56px;display:flex;align-items:center;'
        f'justify-content:center;font-size:28px">{icon}</div>'
    )
    tip = f' title="{tooltip}"' if tooltip else ""
    return (
        f'<div{tip} style="background:#FFFFFF;border:1px solid rgba(15,56,90,0.10);'
        f'border-left:4px solid {color};border-radius:12px;'
        f'padding:14px 16px;display:flex;align-items:center;gap:12px;min-height:84px;'
        f'box-shadow:0 2px 8px rgba(15,56,90,0.07);cursor:help">'
        f'<div style="flex-shrink:0">{arc}</div>'
        f'<div style="flex:1;min-width:0">'
        f'<div style="font-size:10px;color:#6a8a9e;text-transform:uppercase;'
        f'letter-spacing:.5px;margin-bottom:3px">{label}</div>'
        f'<div style="font-size:26px;font-weight:700;color:{color};line-height:1.1">{val}</div>'
        f'<div style="font-size:10px;color:#8aabb0;margin-top:2px">{sub}</div>'
        f'</div></div>'
    )

# ── Helper: tarjeta SVG donut + stats integrados ──────────────────────────────
def _donut_card(proc, pct, done, inp, nst, na_val, color, size=128, r=44, sw=13):
    circ  = 2 * math.pi * r
    cx    = size // 2
    total = max(done + inp + nst + na_val, 1)
    _tip  = (f"Proceso: {proc} · Avance promedio: {pct}% | "
             f"De {done+inp+nst+na_val} programas: "
             f"{done} Finalizados · {inp} En Proceso · {nst} Sin iniciar · {na_val} N/A. "
             f"El % es el promedio del avance de este proceso en todos los programas filtrados.")
    segs  = [(done, "#A6CE38"), (inp, "#1FB2DE"), (nst, "#EC0677"), (na_val, "#c8d8e0")]
    arcs  = f'<circle cx="{cx}" cy="{cx}" r="{r}" fill="none" stroke="#edf1f5" stroke-width="{sw}"/>'
    off   = 0.0
    for cnt, sc in segs:
        if cnt == 0:
            continue
        dash = circ * cnt / total
        arcs += (
            f'<circle cx="{cx}" cy="{cx}" r="{r}" fill="none" stroke="{sc}" '
            f'stroke-width="{sw}" stroke-dasharray="{dash:.3f} {circ:.3f}" '
            f'stroke-dashoffset="-{off:.3f}" transform="rotate(-90 {cx} {cx})"/>'
        )
        off += dash
    arcs += (
        f'<text x="{cx}" y="{cx}" text-anchor="middle" dominant-baseline="central" '
        f'font-size="17" font-weight="bold" font-family="Segoe UI,sans-serif" fill="{color}">{pct}%</text>'
    )
    svg   = f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">{arcs}</svg>'
    label = proc_short_map.get(proc, proc)
    return (
        f'<div title="{_tip}" style="background:#FFFFFF;border:1px solid rgba(15,56,90,.10);'
        f'border-top:3px solid {color};border-radius:10px;'
        f'padding:10px 8px 8px;box-shadow:0 2px 8px rgba(15,56,90,.06);text-align:center;cursor:help">'
        f'<div style="font-size:10px;font-weight:700;color:{color};text-transform:uppercase;'
        f'letter-spacing:.4px;margin-bottom:4px;overflow:hidden;text-overflow:ellipsis;'
        f'white-space:nowrap" title="{proc}">{label}</div>'
        f'<div style="display:flex;justify-content:center;margin:0">{svg}</div>'
        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;margin-top:5px">'
        f'<div style="font-size:9px;color:#5a7a2e;background:#f0f8e8;padding:2px 4px;border-radius:3px;text-align:center"><b>{done}</b> Fin.</div>'
        f'<div style="font-size:9px;color:#0a6a8e;background:#e8f6fc;padding:2px 4px;border-radius:3px;text-align:center"><b>{inp}</b> Proc.</div>'
        f'<div style="font-size:9px;color:#9a0050;background:#fce8f2;padding:2px 4px;border-radius:3px;text-align:center"><b>{nst}</b> Sin ini.</div>'
        f'<div style="font-size:9px;color:#6a8a9e;background:#f0f4f8;padding:2px 4px;border-radius:3px;text-align:center"><b>{na_val}</b> N/A</div>'
        f'</div></div>'
    )

# ── Colores de clasificación ────────────────────────────────────────────────────
CLASIF_COLORS = {
    "Urgente":        ("#EC0677", "#fce8f2"),
    "Prioritario":    ("#FBAF17", "#fdf8e8"),
    "En seguimiento": ("#2980B9", "#EBF5FB"),
    "En curso":       ("#A6CE38", "#f0f8e8"),
}

# ── Ventana emergente: definición de estados ────────────────────────────────────
_dialog_deco = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)

@_dialog_deco("Clasificación de programas · Criterios de prioridad", width="large")
def _dialog_estados():
    st.markdown(
        "Los programas se clasifican automáticamente en **4 niveles de prioridad** "
        "combinando el **periodo de implementación** y el **% de avance general**. "
        "Esta clasificación determina el nivel de atención requerido."
    )
    clasificaciones = [
        ("Urgente",        "#EC0677", "#fce8f2",
         "Periodo 2026-2 o ya en oferta", "Avance general < 70 %",
         "El programa debe implementarse en el próximo semestre (2026-2) pero "
         "aún no alcanza el avance mínimo requerido. Requiere intervención inmediata "
         "de todas las áreas involucradas para evitar incumplimiento."),
        ("Prioritario",    "#FBAF17", "#fdf8e8",
         "Periodo 2027-1", "Avance general < 40 %",
         "El programa implementa en 2027-1 pero registra un avance muy bajo. "
         "Necesita un plan de aceleración antes de que cierre el semestre actual "
         "para no pasar a estado Urgente."),
        ("En seguimiento", "#2980B9", "#EBF5FB",
         "Cualquier periodo", "Avance general < 70 %",
         "El avance es insuficiente para garantizar la implementación a tiempo, "
         "pero el plazo aún permite corrección. Requiere monitoreo activo y reporte "
         "periódico de avance."),
        ("En curso",       "#A6CE38", "#f0f8e8",
         "Cualquier periodo", "Avance general ≥ 70 %",
         "El programa registra un avance sólido en todos los procesos. Se considera "
         "que avanza al ritmo esperado y no presenta riesgo de incumplimiento en "
         "el corto plazo."),
    ]
    for nombre, fg, bg, periodo, avance, desc in clasificaciones:
        st.markdown(
            f'<div style="background:{bg};border-left:5px solid {fg};border-radius:8px;'
            f'padding:12px 16px;margin-bottom:10px">'
            f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">'
            f'<span style="background:{fg};color:white;font-weight:700;font-size:12px;'
            f'padding:3px 12px;border-radius:12px">{nombre}</span>'
            f'<span style="font-size:11px;color:#4a6a7e">📅 {periodo} &nbsp;·&nbsp; 📊 {avance}</span>'
            f'</div>'
            f'<div style="font-size:12px;color:#2a4a5e;line-height:1.5">{desc}</div></div>',
            unsafe_allow_html=True,
        )
    st.markdown(
        '<div style="background:#f0f4f8;border-radius:8px;padding:10px 14px;margin-top:4px">'
        '<span style="font-size:11px;color:#4a6a7e">ℹ️ La clasificación se recalcula '
        'automáticamente al aplicar los filtros de modalidad, facultad o periodo.</span></div>',
        unsafe_allow_html=True,
    )
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    if st.button("Cerrar", use_container_width=True, type="secondary"):
        st.rerun()

def _style_clasif_cell(val):
    """CSS para celda de clasificación."""
    if val in CLASIF_COLORS:
        fg, bg = CLASIF_COLORS[val]
        return f"background-color:{bg};color:{fg};font-weight:700;text-align:center"
    return ""

def _style_avance_cell(val):
    """CSS para celda de avance %."""
    try:
        pct = int(str(val).replace("%", "").strip())
    except Exception:
        return ""
    if pct >= 70:
        return "background-color:#f0f8e8;color:#5a7a2e;font-weight:700;text-align:center"
    if pct >= 40:
        return "background-color:#fef9e8;color:#8a6000;font-weight:700;text-align:center"
    return "background-color:#fce8f2;color:#9a0050;font-weight:700;text-align:center"

def _excel_bytes(df_export):
    """Genera Excel formateado con openpyxl."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    clasif_hex = {
        "Urgente":        ("EC0677", "FFFFFF"),
        "Prioritario":    ("FBAF17", "FFFFFF"),
        "En seguimiento": ("2980B9", "FFFFFF"),
        "En curso":       ("A6CE38", "FFFFFF"),
    }

    def _fill(hex6):
        return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    headers = list(df_export.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if h in PROCESO_COLOR:
            cell.fill = _fill(PROCESO_COLOR[h].lstrip("#"))
        elif h == "Prioridad":
            cell.fill = _fill("0F385A")
        elif h == "Avance %":
            cell.fill = _fill("1FB2DE")
        else:
            cell.fill = _fill("0F385A")
    ws.row_dimensions[1].height = 36

    for ri, row in enumerate(df_export.itertuples(index=False), 2):
        for ci, (h, val) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left" if ci <= 4 else "center", vertical="center")
            if h == "Prioridad" and str(val) in clasif_hex:
                bg, fg = clasif_hex[str(val)]
                cell.fill = _fill(bg)
                cell.font = Font(bold=True, color=fg, size=10)
                cell.alignment = Alignment(horizontal="center")
            elif h == "Avance %":
                try:
                    pct = int(str(val).replace("%", ""))
                    bg = "A6CE38" if pct >= 70 else ("2980B9" if pct >= 40 else "EC0677")
                    fg = "FFFFFF"
                    cell.fill = _fill(bg)
                    cell.font = Font(bold=True, color=fg, size=10)
                    cell.alignment = Alignment(horizontal="center")
                except Exception:
                    pass
            elif h in PROCESO_COLOR:
                hex6 = PROCESO_COLOR[h].lstrip("#")
                r2, g2, b2 = int(hex6[:2], 16), int(hex6[2:4], 16), int(hex6[4:], 16)
                light = f"{min(r2+190,255):02X}{min(g2+190,255):02X}{min(b2+190,255):02X}"
                cell.fill = _fill(light)
                cell.font = Font(color=hex6, bold=True, size=10)
                cell.alignment = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        col_vals = [ws.cell(r, ci).value or "" for r in range(1, ws.max_row + 1)]
        max_w = max(len(str(v)) for v in col_vals) if col_vals else 10
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 3, 35)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Helpers HTML compartidos (usados en tab2 y tab_prio) ───────────────────────
def _p_esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

def _p_icon(val):
    try: v = float(val)
    except Exception: return '<span style="color:#b0bec5;font-size:16px">—</span>'
    import math as _m
    if _m.isnan(v): return '<span style="color:#b0bec5;font-size:13px;font-weight:600">N/A</span>'
    if v >= 100: return '<span style="font-size:16px">✅</span>'
    if v > 0:    return '<span style="font-size:16px">⚠️</span>'
    return              '<span style="font-size:16px">🔴</span>'

def _p_syl(s):
    if s == "Si":  return '<span style="font-size:16px">✅</span>'
    if s == "NO":  return '<span style="font-size:16px">🔴</span>'
    return '<span style="color:#b0bec5;font-size:13px;font-weight:600">N/A</span>'

def _p_bar(pct):
    try:
        pct = float(pct if pct is not None else 0)
    except Exception:
        pct = 0.0
    import math as _m2
    if _m2.isnan(pct): pct = 0.0
    clr = "#15803d" if pct >= 70 else ("#d97706" if pct >= 40 else "#dc2626")
    bar = "#22c55e" if pct >= 70 else ("#f59e0b" if pct >= 40 else "#ef4444")
    return (f'<div style="min-width:70px;text-align:center">'
            f'<div style="font-size:12px;font-weight:700;color:{clr};margin-bottom:3px">{int(pct)}%</div>'
            f'<div style="height:6px;background:#e2e8f0;border-radius:4px;overflow:hidden">'
            f'<div style="width:{min(pct,100):.0f}%;height:100%;background:{bar};border-radius:4px;'
            f'box-shadow:0 1px 2px rgba(0,0,0,.15)"></div>'
            f'</div></div>')

def _p_badge(txt, clr):
    r,g,b = int(clr[1:3],16), int(clr[3:5],16), int(clr[5:7],16)
    return (f'<span style="background:rgba({r},{g},{b},0.12);color:{clr};font-size:10px;'
            f'font-weight:700;padding:3px 9px;border-radius:12px;white-space:nowrap">{_p_esc(txt)}</span>')

def _p_star(av, per, verde_2026=False, amarillo_2026=False):
    if verde_2026:
        return '<span style="color:#22c55e;font-size:14px" title="Puede implementarse">★</span>'
    if amarillo_2026:
        return '<span style="color:#FBAF17;font-size:14px" title="Con esfuerzo podría implementarse">★</span>'
    if "2026" in str(per):
        return '<span style="color:#EC0677;font-size:14px" title="No se podría implementar en 2026-2">★</span>'
    return ''

# ── Tabs principales ────────────────────────────────────────────────────────────
tab0, tab_prio, tab2 = st.tabs(["🏆 Resumen Ejecutivo", "🎯 Priorización", "🏛️ Por Facultad y Programa"])

# ── Tab 0: Resumen Ejecutivo ───────────────────────────────────────────────────
with tab0:
    # ── Filtro dentro del tab ──────────────────────────────────────────────────
    with st.container():
        _lb1, _in1, _sp0, _lb2, _in2, _btn0 = st.columns([0.55, 2.2, 0.05, 0.65, 1.9, 0.65])
        with _lb1: st.markdown(f'<div {_LBL}>📋 MODALIDAD</div>', unsafe_allow_html=True)
        with _in1:
            if _use_pills: st.pills("mod", _mods_ops, selection_mode="multi", key="flt_mod", label_visibility="collapsed")
            else: st.multiselect("mod", _mods_ops, key="flt_mod", label_visibility="collapsed", placeholder="Todas")
        with _lb2: st.markdown(f'<div {_LBL}>🏛️ FACULTAD</div>', unsafe_allow_html=True)
        with _in2:
            if _use_pills: st.pills("fac", fac_ops, selection_mode="multi", key="flt_fac", label_visibility="collapsed")
            else: st.multiselect("fac", fac_ops, key="flt_fac", label_visibility="collapsed", placeholder="Todas")
        with _btn0: st.button("✕ LIMPIAR", on_click=_clear_app, type="primary", key="app_clear")
        _lb3, _in3, _sp0b, _lb_nivel, _in_nivel, _cnt0 = st.columns([0.55, 2.2, 0.05, 0.65, 1.9, 0.65])
        with _lb3: st.markdown(f'<div {_LBL}>📅 PERÍODO</div>', unsafe_allow_html=True)
        with _in3:
            if _use_pills: st.pills("per", _pers_ops, selection_mode="multi", key="flt_per", label_visibility="collapsed")
            else: st.multiselect("per", _pers_ops, key="flt_per", label_visibility="collapsed", placeholder="Todos")
        with _lb_nivel: st.markdown(f'<div {_LBL}>🎓 NIVEL</div>', unsafe_allow_html=True)
        with _in_nivel:
            if _use_pills: st.pills("nivel", _niveles_ops, selection_mode="multi", key="flt_nivel", label_visibility="collapsed")
            else: st.multiselect("nivel", _niveles_ops, key="flt_nivel", label_visibility="collapsed", placeholder="Todos")
        with _cnt0:
            st.markdown(f'<div style="padding-top:9px;font-size:12px;color:#6a8a9e;text-align:right">'
                        f'Mostrando <b style="color:#0F385A">{n}</b> de '
                        f'<b style="color:#0F385A">{len(df_raw)}</b></div>', unsafe_allow_html=True)

    # ── CSS propio del Resumen Ejecutivo (estilos del HTML de referencia) ──────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700;800&family=Barlow+Condensed:wght@600;700;800&display=swap');
    .re-scard{background:#fff;border-radius:10px;padding:14px 16px;border-left:4px solid;
      box-shadow:0 1px 4px rgba(0,0,0,.07);margin-bottom:4px;}
    .re-snum{font-family:'Barlow Condensed',sans-serif;font-size:2.2rem;font-weight:800;
      line-height:1;color:#0F385A;}
    .re-slbl{font-size:.71rem;color:#475569;font-weight:700;text-transform:uppercase;
      letter-spacing:.06em;margin-top:3px;}
    .re-sec{font-family:'Barlow Condensed',sans-serif;font-size:1rem;font-weight:800;
      text-transform:uppercase;letter-spacing:.06em;display:flex;align-items:center;
      gap:8px;margin:18px 0 10px;color:#0F385A;border-bottom:2px solid #e2e8f0;padding-bottom:6px;}
    .re-rcard{background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);
      overflow:hidden;margin-bottom:8px;display:flex;flex-direction:column;height:100%;}
    .re-rcard-hdr{padding:9px 12px 7px;display:flex;align-items:flex-start;gap:8px;}
    .re-rtbl{width:100%;border-collapse:collapse;font-size:.74rem;}
    .re-rtbl thead th{background:#f1f5f9;padding:5px 8px;text-align:left;font-weight:700;
      font-size:.63rem;text-transform:uppercase;letter-spacing:.05em;color:#475569;
      border-bottom:1px solid #e2e8f0;position:sticky;top:0;z-index:1;}
    .re-rtbl tbody tr:nth-child(even){background:#f8fafc;}
    .re-rtbl tbody tr:hover{background:#e0f4fb;}
    .re-rtbl td{padding:5px 8px;border-bottom:1px solid #f1f5f9;vertical-align:middle;}
    .re-rtbl-wrap{flex:1;overflow-x:auto;overflow-y:auto;max-height:320px;}
    .re-pbar{display:inline-flex;align-items:center;gap:4px;}
    .re-pbar-w{width:50px;height:5px;background:#e2e8f0;border-radius:99px;overflow:hidden;display:inline-block;vertical-align:middle;}
    .re-pbar-f{height:100%;border-radius:99px;}
    .re-ecard{background:#fff;border-radius:10px;padding:14px 16px;
      box-shadow:0 1px 4px rgba(0,0,0,.07);border-left:4px solid;}
    .re-estats{display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:9px;}
    .re-estat{display:flex;flex-direction:column;align-items:center;padding:6px 4px;border-radius:7px;}
    .re-estat-num{font-family:'Barlow Condensed',sans-serif;font-size:1.5rem;font-weight:800;line-height:1;}
    .re-estat-lbl{font-size:.59rem;font-weight:700;text-transform:uppercase;letter-spacing:.03em;margin-top:2px;}
    .re-ef{background:#dcfce7;}.re-ef .re-estat-num,.re-ef .re-estat-lbl{color:#15803d;}
    .re-ep{background:#dbeafe;}.re-ep .re-estat-num,.re-ep .re-estat-lbl{color:#1d4ed8;}
    .re-es{background:#fee2e2;}.re-es .re-estat-num,.re-es .re-estat-lbl{color:#dc2626;}
    .re-ena{background:#f1f5f9;}.re-ena .re-estat-num,.re-ena .re-estat-lbl{color:#64748b;}
    .re-ebar-row{display:flex;align-items:center;gap:8px;}
    .re-ebar-w{flex:1;height:6px;background:#e2e8f0;border-radius:99px;overflow:hidden;}
    .re-ebar-f{height:100%;border-radius:99px;}
    .re-ebar-pct{font-family:'Barlow Condensed',sans-serif;font-size:.95rem;font-weight:800;
      color:#0F385A;min-width:34px;text-align:right;}
    .re-ok{padding:10px 14px;font-size:.78rem;color:#15803d;font-weight:600;background:#f0fdf4;
      border-radius:0 0 10px 10px;}
    </style>
    """, unsafe_allow_html=True)

    # ── SECCIÓN 1: Distribución por modalidad y período ────────────────────────
    _cnt_2027_2 = int(df["PERIODO DE IMPLEMENTACIÓN"].str.contains("2027-2", na=False).sum())
    _cnt_oferta = int(df["PERIODO DE IMPLEMENTACIÓN"].str.contains("oferta", case=False, na=False).sum())
    _mod_lower  = df["MODALIDAD"].str.lower().str.strip()
    _cnt_virt   = int((_mod_lower == "virtual").sum())
    _cnt_pres   = int((_mod_lower == "presencial").sum())
    _cnt_hibr   = int((_mod_lower.isin(["híbrido", "hibrido"])).sum())

    def _re_scard(num, lbl, border_color, lbl_color=None):
        lc = lbl_color or border_color
        return (
            f'<div class="re-scard" style="border-color:{border_color}">'
            f'<div class="re-snum">{num}</div>'
            f'<div class="re-slbl" style="color:{lc}">{lbl}</div></div>'
        )

    def _re_sec_lbl(txt):
        return (f'<div style="font-size:9px;font-weight:700;letter-spacing:.08em;color:#94a3b8;'
                f'text-transform:uppercase;margin:6px 0 3px;padding-left:2px">{txt}</div>')

    st.markdown(_re_sec_lbl("Distribución por Modalidad"), unsafe_allow_html=True)
    sm1, sm2, sm3, sm4 = st.columns(4)
    sm1.markdown(_re_scard(n,          "Total",      "#0F385A", "#0F385A"), unsafe_allow_html=True)
    sm2.markdown(_re_scard(_cnt_virt,  "Virtual",    "#1FB2DE", "#1FB2DE"), unsafe_allow_html=True)
    sm3.markdown(_re_scard(_cnt_pres,  "Presencial", "#A6CE38", "#A6CE38"), unsafe_allow_html=True)
    sm4.markdown(_re_scard(_cnt_hibr,  "Híbrido",    "#FBAF17", "#FBAF17"), unsafe_allow_html=True)

    st.markdown(_re_sec_lbl("Distribución por Período"), unsafe_allow_html=True)
    sp1, sp2, sp3, sp4 = st.columns(4)
    sp1.markdown(_re_scard(cnt_2026,    "2026-2",       "#7c3aed", "#7c3aed"), unsafe_allow_html=True)
    sp2.markdown(_re_scard(cnt_2027_1,  "2027-1",       "#EC0677", "#EC0677"), unsafe_allow_html=True)
    sp3.markdown(_re_scard(_cnt_2027_2, "2027-2",       "#0891b2", "#0891b2"), unsafe_allow_html=True)
    sp4.markdown(_re_scard(_cnt_oferta, "Ya en Oferta", "#A6CE38", "#A6CE38"), unsafe_allow_html=True)

    # ── SECCIÓN 2: Riesgos ────────────────────────────────────────────────────
    st.markdown('<div class="re-sec">⚠️ Riesgos Activos</div>', unsafe_allow_html=True)

    # ── Helpers de riesgo ───────────────────────────────────────────────────────
    _PER_STYLE = {
        "2026-2": ("#dc2626", "#fef2f2"),
        "2027-1": ("#d97706", "#fffbeb"),
        "2027-2": ("#2563eb", "#eff6ff"),
    }

    def _per_badge(per_val):
        clr, bg = _PER_STYLE.get(str(per_val), ("#64748b", "#f1f5f9"))
        return (f'<span style="background:{bg};color:{clr};border:1px solid {clr};'
                f'border-radius:12px;padding:2px 9px;font-size:11px;font-weight:700;'
                f'white-space:nowrap">{per_val}</span>')

    def _av_badge(av):
        clr = "#15803d" if av >= 70 else ("#d97706" if av >= 30 else "#dc2626")
        return f'<span style="font-weight:700;color:{clr};font-size:13px">{av}%</span>'

    def _mod_badge(mod):
        return (f'<span style="background:#e0f2fe;color:#0369a1;border-radius:10px;'
                f'padding:2px 8px;font-size:11px;font-weight:600">{mod}</span>')

    def _pbar_html(pct, color):
        pct = min(max(float(pct or 0), 0), 100)
        clr = "#A6CE38" if pct >= 70 else ("#FBAF17" if pct >= 30 else "#EC0677")
        return (f'<div class="re-pbar">'
                f'<span style="font-weight:700;color:{clr};min-width:30px">{int(pct)}%</span>'
                f'<div class="re-pbar-w"><div class="re-pbar-f" style="width:{pct}%;background:{clr}"></div></div>'
                f'</div>')

    _FAC_CLR = {"FSCC": "#EC0677", "FIDI": "#1FB2DE", "FNGS": "#A6CE38"}

    def _r_rows(mask_df, extra):
        rows = []
        for _, r in mask_df.iterrows():
            fac_s = fac_abrev.get(r.get("FACULTAD", ""), "")
            fac_clr = _FAC_CLR.get(fac_s, "#6a8a9e")
            prog = r["NOMBRE DEL PROGRAMA"]
            prog_html = (
                f'<span style="font-weight:600;color:#0F385A;font-size:12px">{prog}</span>'
                + (f'<br><span style="font-size:10px;font-weight:700;color:{fac_clr}">{fac_s}</span>' if fac_s else "")
            )
            per = str(r.get("PERIODO DE IMPLEMENTACIÓN", "—")).strip()
            row_data = {
                "Programa":  prog_html,
                "Modalidad": _mod_badge(r.get("MODALIDAD", "—")),
                "Período":   _per_badge(per),
            }
            for k, fn in extra.items():
                row_data[k] = fn(r)
            rows.append(row_data)
        return rows

    def _rpct(pct):
        """Porcentaje coloreado compacto para tablas de riesgo."""
        v = min(max(float(pct or 0), 0), 100)
        clr = "#15803d" if v >= 90 else ("#d97706" if v >= 70 else "#dc2626")
        disp = f"{v:.1f}%" if v % 1 != 0 else f"{int(v)}%"
        return f'<span style="font-weight:700;color:{clr};font-size:12px">{disp}</span>'

    def _render_rcard(title, desc, color, rows, show_cols, icon="⚠️",
                      tbl_max_height="320px", card_min_height=None):
        """Renderiza un risk-card compacto con cabecera visual y tabla scrollable."""
        n_r = len(rows)
        badge_color = color if n_r > 0 else "#15803d"
        r, g, b = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
        hdr_bg = f"rgba({r},{g},{b},0.06)"
        hdr_bd = f"rgba({r},{g},{b},0.18)"
        card_style = f"border-top:3px solid {badge_color}"
        if card_min_height:
            card_style += f";min-height:{card_min_height}"
        hdr = (
            f'<div class="re-rcard" style="{card_style}">'
            f'<div style="background:{hdr_bg};padding:9px 12px 7px;'
            f'border-bottom:1px solid {hdr_bd}">'
            f'<div style="display:flex;align-items:flex-start;'
            f'justify-content:space-between;gap:6px">'
            f'<div style="flex:1">'
            f'<div style="font-size:10px;font-weight:800;color:{badge_color};'
            f'text-transform:uppercase;letter-spacing:.04em;line-height:1.3">'
            f'{icon}&nbsp;{title}</div>'
            f'<div style="font-size:8.5px;color:#64748b;margin-top:2px;line-height:1.4">'
            f'{desc}</div></div>'
            f'<div style="font-family:\'Barlow Condensed\',sans-serif;font-size:1.8rem;'
            f'font-weight:800;color:{badge_color};line-height:1;flex-shrink:0;'
            f'padding-left:8px">{n_r}</div>'
            f'</div></div>'
        )
        if n_r == 0:
            return hdr + '<div class="re-ok">✅ Sin programas en este riesgo</div></div>'
        tbl_style = f"max-height:{tbl_max_height}" if tbl_max_height != "none" else "max-height:none"
        thead = (f'<div class="re-rtbl-wrap" style="overflow-x:auto;overflow-y:auto;{tbl_style}">'
                 f'<table class="re-rtbl"><thead><tr>')
        for c in show_cols:
            align = "left" if c == "Programa" else "center"
            thead += f'<th style="text-align:{align}">{c}</th>'
        thead += '</tr></thead><tbody>'
        tbody = ""
        for row in rows:
            tbody += "<tr>"
            for c in show_cols:
                align = "left" if c == "Programa" else "center"
                tbody += f'<td style="text-align:{align}">{row.get(c, "—")}</td>'
            tbody += "</tr>"
        return hdr + f'{thead}{tbody}</tbody></table></div></div>'

    PERIODO_ORDER = {"2026-2": 0, "2027-1": 1, "2027-2": 2}

    df_risk = df.copy()

    _pc  = df_risk["pc_pct"]  if "pc_pct"  in df_risk.columns else pd.Series([0.0]*len(df_risk), index=df_risk.index)
    _cf  = df_risk["cf_st"]   if "cf_st"   in df_risk.columns else pd.Series(["nostart"]*len(df_risk), index=df_risk.index)
    _pcs = df_risk["pc_st"]   if "pc_st"   in df_risk.columns else pd.Series(["nostart"]*len(df_risk), index=df_risk.index)
    _ban = df_risk["ban_pct"] if "ban_pct" in df_risk.columns else pd.Series([0.0]*len(df_risk), index=df_risk.index)
    _con = df_risk["conv_pct"]if "conv_pct"in df_risk.columns else pd.Series([0.0]*len(df_risk), index=df_risk.index)
    _syl = df_risk["syl_val"] if "syl_val" in df_risk.columns else pd.Series(["Si"]*len(df_risk), index=df_risk.index)
    _per = df_risk["PERIODO DE IMPLEMENTACIÓN"].str.strip() if "PERIODO DE IMPLEMENTACIÓN" in df_risk.columns else pd.Series([""]*len(df_risk), index=df_risk.index)

    def _sort_risk(risk_df):
        per_col = "PERIODO DE IMPLEMENTACIÓN"
        if per_col in risk_df.columns:
            tmp = risk_df.copy()
            tmp["_por"] = tmp[per_col].str.strip().map(PERIODO_ORDER).fillna(99)
            return tmp.sort_values(["_por", "avance_general"], ascending=[True, False]).drop(columns=["_por"])
        return risk_df.sort_values("avance_general", ascending=False)

    # R1 — Producción virtual sin aval financiero
    # Usa df completo (no df_risk) para incluir programas con trámite ministerial
    _pc_all = df["pc_pct"] if "pc_pct" in df.columns else pd.Series([0.0]*len(df), index=df.index)
    _cf_all = df["cf_st"]  if "cf_st"  in df.columns else pd.Series(["nostart"]*len(df), index=df.index)
    r1_df   = _sort_risk(df[(_pc_all > 0) & (_cf_all == "nostart")])
    r1_rows = _r_rows(r1_df, {
        "% Contenidos": lambda r: _rpct(r.get("pc_pct", 0)),
    })

    # R2 — Lanzamiento 2026-2 con contenidos incompletos
    r2_df   = df_risk[(_per == "2026-2") & (_pcs != "na") & (_pc < 100)].copy()
    r2_df   = r2_df.sort_values("pc_pct", ascending=True)
    r2_rows = _r_rows(r2_df, {
        "% Contenidos": lambda r: _rpct(r.get("pc_pct", 0)),
    })

    # R3 — Parametrización en Banner sin producción de contenidos
    r3_df   = _sort_risk(df_risk[(_ban > 0) & (_pc > 0) & (_pc < 100)])
    r3_rows = _r_rows(r3_df, {
        "% Banner": lambda r: _rpct(r.get("ban_pct", 0)),
    })

    # R4 — Avance en producción — Syllabus incompleto
    r4_df   = df_risk[
        (df_risk["MODALIDAD"].str.lower().str.strip().isin(["virtual", "híbrido", "hibrido"])) &
        (_pc > 0) &
        (_syl == "NO")
    ].copy()
    r4_df   = r4_df.sort_values("pc_pct", ascending=False)
    r4_rows = _r_rows(r4_df, {
        "Estado Syllabus": lambda r: ('<span style="background:#fee2e2;color:#dc2626;font-size:10px;'
                                      'font-weight:700;padding:2px 7px;border-radius:8px">NO</span>'),
        "% Contenidos":    lambda r: _rpct(r.get("pc_pct", 0)),
    })

    # R5 — Banner con avance pero convenios incompletos
    _per_r5   = df_risk["PERIODO DE IMPLEMENTACIÓN"].str.strip()
    _conv_raw_idx = next((i for i,(p,_,_,t) in enumerate(ETAPAS_MAP) if p=="Convenios Institucionales" and t=="pct"), None)
    _conv_raw_col = f"val_{_conv_raw_idx}" if _conv_raw_idx is not None else None
    _conv_val = (df_risk[_conv_raw_col].str.lower().str.strip()
                 if _conv_raw_col and _conv_raw_col in df_risk.columns
                 else pd.Series([""]*len(df_risk), index=df_risk.index))
    r5_df   = df_risk[
        (_ban > 0) & (_con < 100) &
        (_per_r5 != "Ya está en oferta") &
        (_conv_val != "no aplica")
    ].copy()
    r5_df   = r5_df.sort_values("conv_pct", ascending=True)
    r5_rows = _r_rows(r5_df, {
        "% Avance Convenios": lambda r: _rpct(r.get("conv_pct", 0)),
        "% Avance Banner":    lambda r: _rpct(r.get("ban_pct", 0)),
    })

    rr1, rr2, rr3 = st.columns(3)
    with rr1:
        st.markdown(_render_rcard(
            "Producción virtual sin aval financiero",
            "Virtual · % contenidos > 0 y sin concepto financiero",
            "#dc2626", r1_rows,
            ["Programa", "Período", "% Contenidos"], "⚠️",
            tbl_max_height="none", card_min_height="420px"), unsafe_allow_html=True)
    with rr2:
        st.markdown(_render_rcard(
            "Lanzamiento 2026-2 con contenidos incompletos",
            "Período 2026-2 · Producción < 100% · Menor avance primero",
            "#d97706", r2_rows,
            ["Programa", "Modalidad", "% Contenidos"], "🔔",
            tbl_max_height="none"), unsafe_allow_html=True)
    with rr3:
        st.markdown(_render_rcard(
            "Parametrización en Banner sin producción de contenidos",
            "Banner > 0% y Contenidos < 100%",
            "#7c3aed", r3_rows,
            ["Programa", "Período", "% Banner"], "⚙️",
            tbl_max_height="none", card_min_height="420px"), unsafe_allow_html=True)
    rr4, rr5, rr6 = st.columns(3)
    with rr4:
        st.markdown(_render_rcard(
            "Avance en producción — Syllabus incompleto",
            "Virtual/Híbrido · AK > 0% y Syllabus = NO · Mayor avance primero",
            "#0d9488", r4_rows,
            ["Programa", "Estado Syllabus", "% Contenidos"], "📋",
            tbl_max_height="none"), unsafe_allow_html=True)
    with rr5:
        st.markdown(_render_rcard(
            "Banner con avance sin trámite de convenios",
            "BB > 0% y AS < 100% · Menor avance en convenios primero",
            "#2563eb", r5_rows,
            ["Programa", "% Avance Convenios", "% Avance Banner"], "🤝",
            tbl_max_height="none"), unsafe_allow_html=True)

    # ── SECCIÓN 3: Resumen por Etapa ──────────────────────────────────────────
    st.markdown('<div class="re-sec">📋 Estado por Etapa</div>', unsafe_allow_html=True)

    ETAPA_INFO = [
        ("Gestión Académica",                       "#0F385A"),
        ("Gestión Financiera",                      "#FBAF17"),
        ("Aseguramiento de la Calidad",             "#EC0677"),
        ("Syllabus",                                "#9333ea"),
        ("Producción de Contenidos",                "#A6CE38"),
        ("Convenios Institucionales",               "#42B0B5"),
        ("Parametrizar Reforma en Banner",          "#5C89B5"),
        ("Publicación en Página Web",               "#F47B20"),
    ]

    SHORT_NAME = {
        "Gestión Académica":                       "Gestión Académica",
        "Gestión Financiera":                      "Concepto Financiero",
        "Aseguramiento de la Calidad":             "Aseguramiento Calidad",
        "Syllabus":                                "Syllabus",
        "Producción de Contenidos":                "Producción Contenidos",
        "Convenios Institucionales":               "Convenios",
        "Parametrizar Reforma en Banner":          "Parametrización Banner",
        "Publicación en Página Web":               "Publicación Web",
    }

    etapa_cols_row1 = st.columns(4)
    etapa_cols_row2 = st.columns(4)
    all_etapa_cols = etapa_cols_row1 + etapa_cols_row2

    for ei, (proc, color) in enumerate(ETAPA_INFO):
        p_vals  = df[f"proc_{proc}"]
        done_e  = int((p_vals == 100).sum())
        inp_e   = int(((p_vals > 0) & (p_vals < 100)).sum())
        nst_e   = int((p_vals == 0).sum())
        na_e    = int(p_vals.isna().sum())
        total_e = max(done_e + inp_e + nst_e, 1)
        pct_e   = round(done_e / total_e * 100)
        short  = SHORT_NAME.get(proc, proc)

        card_html = (
            f'<div class="re-ecard" style="border-color:{color}">'
            f'<div style="font-family:\'Barlow Condensed\',sans-serif;font-size:.82rem;'
            f'font-weight:800;color:#0F385A;text-transform:uppercase;letter-spacing:.04em;'
            f'margin-bottom:10px">{short}</div>'
            f'<div class="re-estats">'
            f'<div class="re-estat re-ef"><div class="re-estat-num">{done_e}</div>'
            f'<div class="re-estat-lbl">Final.</div></div>'
            f'<div class="re-estat re-ep"><div class="re-estat-num">{inp_e}</div>'
            f'<div class="re-estat-lbl">Proceso</div></div>'
            f'<div class="re-estat re-es"><div class="re-estat-num">{nst_e}</div>'
            f'<div class="re-estat-lbl">Sin ini.</div></div>'
            f'<div class="re-estat re-ena"><div class="re-estat-num">{na_e}</div>'
            f'<div class="re-estat-lbl">N/A</div></div>'
            f'</div>'
            f'<div class="re-ebar-row">'
            f'<div class="re-ebar-w"><div class="re-ebar-f" style="width:{pct_e}%;background:{color}"></div></div>'
            f'<div class="re-ebar-pct">{pct_e}%</div>'
            f'</div></div>'
        )
        all_etapa_cols[ei].markdown(card_html, unsafe_allow_html=True)

# ── Tab 2: Por facultad ────────────────────────────────────────────────────────
with tab2:
    # ── KPI cards por facultad ─────────────────────────────────────────────────
    _FAC_DEFS = [
        ("Sociedad, Cultura y Creatividad",    "FSCC", "#EC0677"),
        ("Ingeniería, Diseño e Innovación",    "FIDI", "#1FB2DE"),
        ("Negocios, Gestión y Sostenibilidad", "FNGS", "#A6CE38"),
    ]
    _kpi_cols = st.columns(3)
    for _ki, (_fn, _fab, _fc) in enumerate(_FAC_DEFS):
        _fdf = df[df["FACULTAD"].map(fac_labels) == _fn]
        _ftot  = len(_fdf)
        _favg  = int(_fdf["avance_general"].mean()) if _ftot else 0
        _furg  = int((_fdf["_clasif"] == "Urgente").sum())
        _fpri  = int((_fdf["_clasif"] == "Prioritario").sum())
        _fok   = int((_fdf["_clasif"] == "En curso").sum())
        _fseg  = int((_fdf["_clasif"] == "En seguimiento").sum())
        r, g, b = int(_fc[1:3],16), int(_fc[3:5],16), int(_fc[5:7],16)
        _kpi_cols[_ki].markdown(
            f'<div style="background:#FFFFFF;border:1px solid rgba({r},{g},{b},0.30);'
            f'border-top:4px solid {_fc};border-radius:12px;padding:14px 16px;'
            f'box-shadow:0 2px 10px rgba(15,56,90,0.07);margin-bottom:10px">'
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">'
            f'<div style="width:10px;height:10px;border-radius:50%;background:{_fc}"></div>'
            f'<div style="font-size:11px;font-weight:700;color:{_fc};letter-spacing:.3px">{_fab}</div>'
            f'<div style="font-size:10px;color:#8aabb0;flex:1;white-space:nowrap;overflow:hidden;'
            f'text-overflow:ellipsis">{_fn}</div></div>'
            f'<div style="display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:10px">'
            f'<div><div style="font-size:32px;font-weight:800;color:#0F385A;line-height:1">{_ftot}</div>'
            f'<div style="font-size:9px;color:#8aabb0;text-transform:uppercase;letter-spacing:.4px">Programas</div></div>'
            f'<div style="text-align:right">'
            f'<div style="font-size:22px;font-weight:700;color:{_fc};line-height:1">{_favg}%</div>'
            f'<div style="font-size:9px;color:#8aabb0;text-transform:uppercase;letter-spacing:.4px">Avance prom.</div>'
            f'</div></div>'
            f'<div style="height:5px;background:#f0f4f8;border-radius:3px;overflow:hidden;margin-bottom:10px">'
            f'<div style="width:{_favg}%;height:100%;background:{_fc};border-radius:3px;'
            f'opacity:0.75"></div></div>'
            f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:4px">'
            f'<div style="background:#fce8f2;border-radius:6px;padding:4px 8px;text-align:center">'
            f'<div style="font-size:14px;font-weight:700;color:#EC0677">{_furg}</div>'
            f'<div style="font-size:8px;color:#c0005a;text-transform:uppercase;letter-spacing:.3px">Urgentes</div></div>'
            f'<div style="background:#fef9e8;border-radius:6px;padding:4px 8px;text-align:center">'
            f'<div style="font-size:14px;font-weight:700;color:#d97706">{_fpri}</div>'
            f'<div style="font-size:8px;color:#a05b00;text-transform:uppercase;letter-spacing:.3px">Prioritarios</div></div>'
            f'<div style="background:#f0f8e8;border-radius:6px;padding:4px 8px;text-align:center">'
            f'<div style="font-size:14px;font-weight:700;color:#5a7a2e">{_fok}</div>'
            f'<div style="font-size:8px;color:#3a5a10;text-transform:uppercase;letter-spacing:.3px">En curso</div></div>'
            f'<div style="background:#e8f6fc;border-radius:6px;padding:4px 8px;text-align:center">'
            f'<div style="font-size:14px;font-weight:700;color:#0a6a8e">{_fseg}</div>'
            f'<div style="font-size:8px;color:#06526e;text-transform:uppercase;letter-spacing:.3px">Seguimiento</div></div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )

    # ── Filtro por facultad ────────────────────────────────────────────────────
    _FAC_ABREV_LIST = ["FSCC", "FIDI", "FNGS"]
    _FAC_CLR2 = {"FSCC": "#EC0677", "FIDI": "#1FB2DE", "FNGS": "#A6CE38"}
    if "t2_fac" not in st.session_state: st.session_state["t2_fac"] = []
    def _clear_t2(): st.session_state["t2_fac"] = []

    _lt2, _it2, _bt2, _ct2 = st.columns([0.5, 2.0, 0.55, 1.5])
    _LBL_T2 = 'style="padding-top:8px;font-size:11px;font-weight:700;color:#0F385A;letter-spacing:.4px;white-space:nowrap"'
    with _lt2: st.markdown(f'<div {_LBL_T2}>🏛️ FACULTAD</div>', unsafe_allow_html=True)
    with _it2: sel_t2fac = st.pills("t2fac", _FAC_ABREV_LIST, selection_mode="multi",
                                     key="t2_fac", label_visibility="collapsed")
    with _bt2: st.button("✕ LIMPIAR", on_click=_clear_t2, use_container_width=True,
                         type="primary", key="t2_clear")
    with _ct2:
        _t2_cnt = st.empty()

    # Preparar datos
    df_t2 = df.copy()
    if sel_t2fac:
        df_t2 = df_t2[df_t2["FACULTAD"].isin([fac_abrev_inv.get(f, f) for f in sel_t2fac])]

    # Sort by PERIODO DE IMPLEMENTACIÓN then program name
    _per_ord = {"Ya está en oferta": 0, "2026-2": 1, "2027-1": 2, "2027-2": 3}
    df_t2 = df_t2.copy()
    df_t2["_per_sort"] = df_t2["PERIODO DE IMPLEMENTACIÓN"].map(_per_ord).fillna(99)
    df_t2 = df_t2.sort_values(["_per_sort", "NOMBRE DEL PROGRAMA"]).reset_index(drop=True)

    _t2_cnt.markdown(
        f'<div style="padding-top:9px;font-size:12px;color:#6a8a9e;text-align:right">'
        f'Mostrando <b style="color:#0F385A">{len(df_t2)}</b> de '
        f'<b style="color:#0F385A">{len(df)}</b></div>', unsafe_allow_html=True)

    # Render HTML table
    _PER_CLR2 = {"Ya está en oferta": "#1FB2DE", "2026-2": "#EC0677", "2027-1": "#d97706", "2027-2": "#2563eb"}
    _MOD_CLR2 = {"Virtual": "#1FB2DE", "Presencial": "#A6CE38", "Híbrido": "#FBAF17"}
    _TH2 = 'style="background:#0F385A;color:#FFFFFF;font-size:10px;font-weight:700;padding:7px 8px;text-align:left;white-space:nowrap;border-right:1px solid rgba(255,255,255,0.10)"'
    _THC2 = 'style="background:#0F385A;color:#FFFFFF;font-size:10px;font-weight:700;padding:7px 8px;text-align:center;white-space:nowrap;border-right:1px solid rgba(255,255,255,0.10)"'

    rows_t2 = []
    for idx2, (_, row2) in enumerate(df_t2.iterrows()):
        rbg2 = "#FFFFFF" if idx2 % 2 == 0 else "#f8fafc"
        _TDL2 = f'style="padding:6px 8px;text-align:left;vertical-align:middle;border-bottom:1px solid #eef3f8;background:{rbg2}"'
        _TDC2 = f'style="padding:6px 8px;text-align:center;vertical-align:middle;border-bottom:1px solid #eef3f8;background:{rbg2}"'
        prog2  = _p_esc(str(row2.get("NOMBRE DEL PROGRAMA","—")))
        fac2   = fac_abrev.get(str(row2.get("FACULTAD","")), "—")
        fac_c2 = _FAC_CLR2.get(fac2, "#9aabb5")
        mod2   = str(row2.get("MODALIDAD","—"))
        mod_c2 = _MOD_CLR2.get(mod2, "#9aabb5")
        niv2   = _p_esc(str(row2.get("NIVEL","—")))
        sed2   = _p_esc(str(row2.get("SEDE","—")))
        per2   = str(row2.get("PERIODO DE IMPLEMENTACIÓN","—"))
        per_c2 = _PER_CLR2.get(per2, "#9aabb5")
        av2    = int(float(row2.get("avance_general",0) or 0))
        av_c2  = "#15803d" if av2 >= 70 else ("#d97706" if av2 >= 40 else "#dc2626")
        def _mk_badge2(txt, clr):
            r2,g2,b2 = int(clr[1:3],16), int(clr[3:5],16), int(clr[5:7],16)
            return (f'<span style="background:rgba({r2},{g2},{b2},0.12);color:{clr};font-size:10px;'
                    f'font-weight:700;padding:2px 8px;border-radius:10px;white-space:nowrap">{txt}</span>')
        rows_t2.append(
            f'<tr>'
            f'<td {_TDL2}><span style="font-size:11px;font-weight:600;color:#0F385A">{prog2}</span>'
            f'<br><span style="font-size:9px;font-weight:700;color:{fac_c2}">{fac2}</span></td>'
            f'<td {_TDC2}>{_mk_badge2(mod2, mod_c2)}</td>'
            f'<td {_TDC2}><span style="font-size:10px;color:#4a6a7e">{niv2}</span></td>'
            f'<td {_TDC2}><span style="font-size:10px;color:#4a6a7e">{sed2}</span></td>'
            f'<td {_TDC2}>{_mk_badge2(per2, per_c2)}</td>'
            f'<td {_TDC2}>'
            f'<div style="font-size:11px;font-weight:700;color:{av_c2};margin-bottom:2px">{av2}%</div>'
            f'<div style="height:5px;background:#e2e8f0;border-radius:3px;overflow:hidden;min-width:50px">'
            f'<div style="width:{min(av2,100)}%;height:100%;background:{av_c2};border-radius:3px"></div>'
            f'</div></td>'
            f'</tr>'
        )

    tabla_t2 = (
        '<div style="overflow-x:auto;border-radius:12px;border:1.5px solid #b5c9d8;'
        'box-shadow:0 2px 12px rgba(15,56,90,.10);background:#fafdff;margin-top:4px">'
        '<table style="width:100%;border-collapse:separate;border-spacing:0;font-family:\'Segoe UI\',sans-serif">'
        '<thead><tr>'
        f'<th {_TH2} style="min-width:200px">Programa</th>'
        f'<th {_THC2}>Modalidad</th>'
        f'<th {_THC2}>Nivel</th>'
        f'<th {_THC2}>Sede</th>'
        f'<th {_THC2}>Período</th>'
        f'<th {_THC2}>Avance</th>'
        '</tr></thead><tbody>'
        + "".join(rows_t2) +
        '</tbody></table></div>'
    )
    _t2h = max(280, min(900, 80 + len(df_t2) * 50))
    _html_comp.html(tabla_t2, height=_t2h, scrolling=True)
    st.markdown(f'<div style="font-size:11px;color:#6a8a9e;margin-top:4px;text-align:right">'
                f'{len(df_t2)} programas · ordenados por período de implementación</div>',
                unsafe_allow_html=True)

# ── Tab Priorización ──────────────────────────────────────────────────────────
with tab_prio:
    _P_FAC_CLR = {"FSCC": "#EC0677", "FIDI": "#1FB2DE", "FNGS": "#A6CE38"}
    _P_MOD_CLR = {"Virtual": "#1FB2DE", "Presencial": "#A6CE38", "Híbrido": "#FBAF17"}
    _P_PER_CLR = {"2026-2": "#EC0677", "2027-1": "#FBAF17", "2027-2": "#2563eb", "Ya está en oferta": "#1FB2DE"}

    # ── Filtro propio del tab ───────────────────────────────────────────────────
    if "prio_mod" not in st.session_state: st.session_state["prio_mod"] = []
    if "prio_per" not in st.session_state: st.session_state["prio_per"] = ["2026-2"]
    if "prio_fac" not in st.session_state: st.session_state["prio_fac"] = []
    if "prio_nivel" not in st.session_state: st.session_state["prio_nivel"] = []

    def _clear_prio():
        st.session_state["prio_mod"]   = []
        st.session_state["prio_per"]   = ["2026-2"]
        st.session_state["prio_fac"]   = []
        st.session_state["prio_nivel"] = []

    _LBL_P = ('style="padding-top:8px;font-size:11px;font-weight:700;color:#0F385A;'
               'letter-spacing:.4px;white-space:nowrap"')
    with st.container():
        _lp1, _ip1, _sp_p, _lp2, _ip2, _btn_p = st.columns([0.55, 2.2, 0.05, 0.65, 1.9, 0.65])
        with _lp1: st.markdown(f'<div {_LBL_P}>📋 MODALIDAD</div>', unsafe_allow_html=True)
        with _ip1: sel_pmod = st.pills("pmod", sorted(df_raw["MODALIDAD"].dropna().unique().tolist()),
                                       selection_mode="multi", key="prio_mod", label_visibility="collapsed")
        with _lp2: st.markdown(f'<div {_LBL_P}>🏛️ FACULTAD</div>', unsafe_allow_html=True)
        with _ip2: sel_pfac = st.pills("pfac", sorted([fac_abrev.get(f,f) for f in df_raw["FACULTAD"].dropna().unique()]),
                                       selection_mode="multi", key="prio_fac", label_visibility="collapsed")
        with _btn_p: st.button("✕ LIMPIAR", on_click=_clear_prio, type="primary", key="prio_clear")
        _lp3, _ip3, _sp_p2, _lp_nivel, _ip_nivel, _cnt_p = st.columns([0.55, 2.2, 0.05, 0.65, 1.9, 0.65])
        with _lp3: st.markdown(f'<div {_LBL_P}>📅 PERÍODO</div>', unsafe_allow_html=True)
        with _ip3: sel_pper = st.pills("pper", sorted(df_raw["PERIODO DE IMPLEMENTACIÓN"].dropna().unique().tolist()),
                                       selection_mode="multi", key="prio_per", label_visibility="collapsed")
        with _lp_nivel: st.markdown(f'<div {_LBL_P}>🎓 NIVEL</div>', unsafe_allow_html=True)
        with _ip_nivel:
            _niv_ops = [n for n in ["Pregrado", "Posgrado"] if "NIVEL_HOMOLOGADO" in df_raw.columns and n in df_raw["NIVEL_HOMOLOGADO"].values]
            if _use_pills: st.pills("pnivel", _niv_ops, selection_mode="multi", key="prio_nivel", label_visibility="collapsed")
            else: st.multiselect("pnivel", _niv_ops, key="prio_nivel", label_visibility="collapsed", placeholder="Todos")
        with _cnt_p: _prio_cnt = st.empty()

    # ── Filtrar datos ───────────────────────────────────────────────────────────
    df_p = df_raw.copy()
    if sel_pmod:  df_p = df_p[df_p["MODALIDAD"].isin(sel_pmod)]
    if sel_pfac:  df_p = df_p[df_p["FACULTAD"].isin([fac_abrev_inv.get(f,f) for f in sel_pfac])]
    if sel_pper:  df_p = df_p[df_p["PERIODO DE IMPLEMENTACIÓN"].isin(sel_pper)]
    _sel_pnivel = list(st.session_state.get("prio_nivel") or [])
    if _sel_pnivel and "NIVEL_HOMOLOGADO" in df_p.columns:
        df_p = df_p[df_p["NIVEL_HOMOLOGADO"].isin(_sel_pnivel)]
    df_p = enrich_df(df_p) if "pc_pct" not in df_p.columns else df_p

    # Excluir "Ya está en oferta"
    df_p = df_p[~df_p["PERIODO DE IMPLEMENTACIÓN"].str.strip().str.lower().str.contains("oferta", na=False)].copy()

    # (programa, modalidad, periodo) → estrella forzada — asignación explícita
    _VERDE_FORZADO = {
        ("Licenciatura en Educación Infantil",        "Virtual",    "2026-2"),
        ("Licenciatura en Educación Básica Primaria", "Virtual",    "2026-2"),
        ("Técnica Profesional Judicial",              "Virtual",    "2026-2"),
        ("Administración de Empresas",                "Presencial", "2026-2"),
        ("Contaduría Pública",                        "Presencial", "2026-2"),
    }
    _AMARILLO_FORZADO = {
        ("Contaduría Pública",     "Virtual", "2026-2"),
        ("Ingeniería de Software", "Virtual", "2026-2"),
        ("Derecho",                "Virtual", "2026-2"),
    }

    def _es_verde(row):
        prog = str(row.get("NOMBRE DEL PROGRAMA", "")).strip()
        per  = str(row.get("PERIODO DE IMPLEMENTACIÓN","")).strip()
        mod  = str(row.get("MODALIDAD", "")).strip()
        if prog == "Tecnología en Gestión Ambiental" and per == "2027-1":
            return True
        return (prog, mod, per) in _VERDE_FORZADO

    def _es_amarillo(row):
        prog = str(row.get("NOMBRE DEL PROGRAMA", "")).strip()
        per  = str(row.get("PERIODO DE IMPLEMENTACIÓN","")).strip()
        mod  = str(row.get("MODALIDAD", "")).strip()
        return (prog, mod, per) in _AMARILLO_FORZADO

    # clasificar
    _PER_ORD = {"2026-2": 0, "2027-1": 1, "2027-2": 2}
    if len(df_p) > 0:
        df_p["_clasif"] = df_p.apply(
            lambda r: clasificar_programa(r["avance_general"], r["PERIODO DE IMPLEMENTACIÓN"]), axis=1)
        df_p["_verde_2026"]   = df_p.apply(_es_verde,   axis=1)
        df_p["_amarillo_2026"] = df_p.apply(_es_amarillo, axis=1)
        def _star_ord(row):
            av  = float(row.get("avance_general", 0) or 0)
            per = str(row.get("PERIODO DE IMPLEMENTACIÓN", "")).strip()
            if row.get("_verde_2026"):    return (0, -av)   # verde
            if row.get("_amarillo_2026"): return (1, -av)   # amarillo (forzado explícito)
            if "2026" in per:             return (2, -av)   # rojo 2026-2
            return (3, -av)                                 # 2027-x
        df_p["_sort_key"] = df_p.apply(lambda r: _star_ord(r), axis=1)
        df_p["_per_ord"]  = df_p["PERIODO DE IMPLEMENTACIÓN"].apply(
            lambda x: _PER_ORD.get(str(x).strip(), 4))
        df_p = df_p.sort_values(["_per_ord", "_sort_key"])

    _prio_cnt.markdown(
        f'<div style="padding-top:9px;font-size:12px;color:#6a8a9e;text-align:right">'
        f'Mostrando <b style="color:#0F385A">{len(df_p)}</b> de '
        f'<b style="color:#0F385A">{len(df_raw)}</b></div>', unsafe_allow_html=True)

    # ── Leyenda ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div style="display:flex;gap:16px;margin:4px 0 8px;flex-wrap:wrap">'
        '<span style="font-size:11px;color:#6a8a9e"><span style="color:#A6CE38;font-size:14px">★</span> Puede implementarse en 2026-2</span>'
        '<span style="font-size:11px;color:#6a8a9e"><span style="color:#FBAF17;font-size:14px">★</span> Con esfuerzo podría implementarse</span>'
        '<span style="font-size:11px;color:#6a8a9e"><span style="color:#EC0677;font-size:14px">★</span> No se podría implementar en 2026-2</span>'
        '</div>', unsafe_allow_html=True)

    _PRIO_ETAPAS = [
        ("G.Académica",  "proc_Gestión Académica",                       "proc"),
        ("C.Financiero", "proc_Gestión Financiera",                       "proc"),
        ("Aseguramient", "proc_Aseguramiento de la Calidad",              "proc"),
        ("Syllabus",     "syl_val",                                        "syl"),
        ("Prod.Cont.",   "pc_pct",                                         "bar"),
        ("Convenios",    "conv_pct",                                      "bar"),
        ("Banner",       "ban_pct",                                        "bar"),
        ("Tipo Trámite", "Tipo de trámite de aseguramiento de la calidad", "tramite"),
        ("Fecha Notif.", "Fecha de\nDocumentos de notificación MEN",       "text"),
        ("Req. Min.",    "¿Requiere aprobación ministerial?",              "text"),
    ]
    _PRIO_CLR = {"Urgente":("#EC0677","#fce8f2"),"Prioritario":("#FBAF17","#fdf8e8"),
                 "En seguimiento":("#2980B9","#EBF5FB"),"En curso":("#A6CE38","#f0f8e8")}

    if len(df_p) == 0:
        st.info("Sin programas para los filtros seleccionados.")
    else:
        TH_P  = ('style="background:#0F385A;color:#FFFFFF;font-size:9px;font-weight:700;'
                 'padding:6px 4px;text-align:center;white-space:nowrap;position:sticky;top:0;z-index:2;'
                 'border-right:1px solid rgba(255,255,255,0.10)"')
        THL_P = ('style="background:#0F385A;color:#FFFFFF;font-size:9px;font-weight:700;'
                 'padding:6px 8px;text-align:left;white-space:nowrap;position:sticky;top:0;z-index:2;'
                 'border-right:1px solid rgba(255,255,255,0.10)"')
        _PER_HDR_CLR = {"2026-2": "#A6CE38", "2027-1": "#FBAF17", "2027-2": "#EC0677"}
        _ncols = 3 + len(_PRIO_ETAPAS)
        rows_p = []
        _cur_per = None
        _row_idx = 0
        for _, row in df_p.iterrows():
            per = str(row.get("PERIODO DE IMPLEMENTACIÓN", "—")).strip()
            # ── Encabezado de grupo por período ──
            if per != _cur_per:
                _cur_per = per
                _ph_clr  = _PER_HDR_CLR.get(per, "#9aabb5")
                _per_cnt = int((df_p["PERIODO DE IMPLEMENTACIÓN"].str.strip() == per).sum())
                rows_p.append(
                    f'<tr><td colspan="{_ncols}" style="background:{_ph_clr};color:#FFFFFF;'
                    f'font-size:11px;font-weight:800;padding:7px 12px;letter-spacing:.3px">'
                    f'{_p_esc(per)} · {_per_cnt} PROGRAMAS</td></tr>'
                )
                _row_idx = 0
            rbg = "#FFFFFF" if _row_idx % 2 == 0 else "#f8fafc"
            _row_idx += 1
            TD  = (f'style="padding:4px 3px;text-align:center;vertical-align:middle;'
                   f'border-bottom:1px solid #eef3f8;background:{rbg}"')
            TDL = (f'style="padding:4px 6px;text-align:left;vertical-align:middle;'
                   f'border-bottom:1px solid #eef3f8;background:{rbg};min-width:140px;max-width:200px"')
            prog   = _p_esc(row.get("NOMBRE DEL PROGRAMA","—"))
            fac_s  = fac_abrev.get(str(row.get("FACULTAD","")), "—")
            fac_c  = _P_FAC_CLR.get(fac_s, "#9aabb5")
            mod    = str(row.get("MODALIDAD","—"))
            mod_c  = _P_MOD_CLR.get(mod, "#9aabb5")
            per_c  = _P_PER_CLR.get(per, "#9aabb5")
            av     = int(float(row.get("avance_general",0) or 0))
            clasif = row.get("_clasif","En curso")
            fg_c, bg_c = _PRIO_CLR.get(clasif, ("#0F385A","#f0f4f8"))

            etapa_cells = []
            for _, col_key, typ in _PRIO_ETAPAS:
                val = row.get(col_key)
                if typ == "proc":
                    try: v = None if val in ("",None) else float(val)
                    except: v = None
                    etapa_cells.append(f'<td {TD}>{_p_icon(v)}</td>')
                elif typ == "syl":
                    etapa_cells.append(f'<td {TD}>{_p_syl(str(val or "N/A"))}</td>')
                elif typ == "bar":
                    if col_key == "pc_pct" and mod == "Presencial":
                        etapa_cells.append(f'<td {TD}><span style="font-size:9px;color:#94a3b8;font-style:italic">No aplica</span></td>')
                    else:
                        try:
                            pct = float(val if val is not None else 0)
                            etapa_cells.append(f'<td {TD}>{_p_bar(pct)}</td>')
                        except Exception:
                            etapa_cells.append(f'<td {TD}><span style="color:#b0bec5;font-size:13px">—</span></td>')
                elif typ == "tramite":
                    tv = val if val not in [None,"","nan"] else "—"
                    etapa_cells.append(f'<td style="padding:5px 4px;text-align:center;vertical-align:middle;'
                                      f'border-bottom:1px solid #eef3f8;background:{rbg};min-width:110px">'
                                      f'<span style="font-size:9px;background:#eff6ff;color:#2563eb;'
                                      f'padding:2px 5px;border-radius:6px;font-weight:600;'
                                      f'display:inline-block;line-height:1.3;word-break:break-word">'
                                      f'{_p_esc(str(tv))}</span></td>')
                else:
                    tv = val if val not in [None,"","nan"] else "—"
                    etapa_cells.append(f'<td style="padding:5px 4px;text-align:center;vertical-align:middle;'
                                      f'border-bottom:1px solid #eef3f8;background:{rbg};min-width:80px">'
                                      f'<span style="font-size:10px;color:#4a6a7e;word-break:break-word;'
                                      f'display:inline-block;line-height:1.3">'
                                      f'{_p_esc(str(tv))}</span></td>')

            rows_p.append(
                f'<tr>'
                f'<td {TDL}>'
                f'<span style="margin-right:3px">{_p_star(av, per, row.get("_verde_2026", False), row.get("_amarillo_2026", False))}</span>'
                f'<span style="font-size:11px;font-weight:600;color:#0F385A">{prog}</span>'
                f'<br><span style="font-size:10px;font-weight:700;color:{fac_c}">{fac_s}</span></td>'
                f'<td {TD}>{_p_badge(mod, mod_c)}</td>'
                f'<td {TD}>{_p_badge(per, per_c)}</td>'
                + "".join(etapa_cells) +
                f'</tr>'
            )

        hdrs = "".join(f'<th {TH_P}>{h}</th>' for h,_,_ in _PRIO_ETAPAS)
        tabla_p = (
            '<div style="overflow-x:auto;border-radius:12px;'
            'border:1.5px solid #b5c9d8;box-shadow:0 2px 12px rgba(15,56,90,.10);background:#fafdff">'
            '<table style="width:100%;table-layout:auto;border-collapse:separate;border-spacing:0;font-family:\'Segoe UI\',sans-serif">'
            '<thead><tr>'
            f'<th {THL_P} style="min-width:140px;max-width:200px">Programa</th>'
            f'<th {TH_P}>Modalidad</th>'
            f'<th {TH_P}>Período</th>'
            + hdrs +
            '</tr></thead><tbody>'
            + "".join(rows_p) +
            '</tbody></table></div>'
        )
        _n_per_grps = len(df_p["PERIODO DE IMPLEMENTACIÓN"].unique())
        _tph = max(300, min(1400, 100 + len(df_p) * 62 + _n_per_grps * 36))
        _html_comp.html(tabla_p, height=_tph, scrolling=True)
        st.markdown(f'<div style="font-size:11px;color:#6a8a9e;margin-top:4px;text-align:right">'
                    f'{len(df_p)} programas · ordenados por prioridad y avance general</div>',
                    unsafe_allow_html=True)

